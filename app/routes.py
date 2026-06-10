from flask import (
    Blueprint, render_template, redirect, url_for, request,
    flash, jsonify, send_file, session
)

# Flask Extensions
from flask_login import (
    login_user, logout_user, login_required, current_user
)

# Database & Extensions
from extensions import db, csrf, migrate
from datetime import datetime, timedelta, date

# Models
from app.models import (
    Product, Customer, Sale, SaleItem, DebtPayment,
    DailyExpense, User, ReturnProduct, Loan, LoanPayment,
    InventoryLoss, Creditor, DebtTransaction
)

# Forms
from app.forms import (
    ProductForm, DeleteForm, CustomerForm, SaleForm,
    DebtPaymentForm, DailyExpenseForm, LoginForm, UserForm,
    ReturnProductForm, EditSaleForm, LoanForm, LoanPaymentForm,
    SaleSearchForm, EditSaleFullForm, InventoryWriteOffForm,
    CreditorForm, DebtTransactionForm, TransactionForm, PaymentForm
)

# Python Standard Library
import csv
import io
import uuid
import socket
import hashlib
from decimal import Decimal
import traceback
from flask import render_template, request, redirect, url_for
from license_core import get_hardware_id, generate_license, get_app_base_path  # 🔴 comment شد چون توابع داخلی داریم
from app.excel_utils import generate_creditor_excel_report
import os
import sys
from app.forms import TransactionForm

from flask import Blueprint, render_template, request, redirect, url_for, session, flash
from flask_login import login_user, logout_user, login_required, current_user
from .extensions import db, login_manager, csrf
from .models import User, Product, Customer, Sale
from .forms import LoginForm
from flask import Blueprint, render_template, redirect, url_for, request, flash
from app.extensions import csrf
from datetime import datetime
from app.models import ForeignProduct
from app.forms import ForeignProductSaleForm

import shutil
from werkzeug.utils import secure_filename
from flask import request
from sqlalchemy.orm import joinedload

from sqlalchemy import func
from .excel_export import export_to_excel, format_currency
from app.notification_manager import NotificationManager

from flask import render_template, request, redirect, url_for, flash
from app.models import Product
from app.extensions import db
main_bp = Blueprint("main_bp", __name__)



from datetime import datetime, timedelta
from flask import (
    Blueprint, render_template, redirect,
    url_for, request, session
)
from app import db
from app.models import License
from app.utils import get_hardware_id  # همان تابع HW ID خودت
from datetime import datetime, timedelta
from app.models import Customer, Creditor, Sale, Product
from datetime import datetime, timedelta
import calendar
from sqlalchemy import func, and_
from flask import request
from app.models import Sale, SaleItem, Product, ForeignProduct, DailyExpense, InventoryLoss
# در بالای routes.py، همراه سایر importها
from app.models import CashBalance, CashTransaction, Notification
from flask_wtf import FlaskForm
from wtforms import FloatField, DateField, SelectField, TextAreaField, StringField, SubmitField
from wtforms.validators import DataRequired, NumberRange, Optional
from app.models import Sale, SaleItem, Customer, Product, CashBalance
    



# =====================================================
# تنظیمات آپلود عکس داشبورد
# =====================================================
import os
import uuid
from werkzeug.utils import secure_filename

DASHBOARD_UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
# ----------------------------
# تولید کلید لایسنس (HW + Time)
# app/routes.py - قبل از همه روت‌ها
# app/routes.py - این بخش را کامل جایگزین کنید
# app/routes.py - نسخه بسیار ساده برای تست

@main_bp.before_request
def block_without_license():
    """بررسی لایسنس - نسخه ساده"""
    from app.models import License
    
    # لیست مسیرهای آزاد
    free_paths = ['/license', '/license/activate', '/license/trial', '/login', '/logout', '/static']
    
    # اگر مسیر در لیست آزاد است
    if request.path in free_paths or request.path.startswith('/static'):
        return None
    
    # اگر کاربر لاگین نیست
    if not current_user.is_authenticated:
        return None
    
    # بررسی لایسنس
    license = License.query.first()
    
    # اگر لایسنس وجود ندارد یا منقضی شده
    if not license:
        flash("⚠️ لطفاً لایسنس را فعال کنید", "warning")
        return redirect('/license')
    
    # بررسی انقضا
    try:
        if not license.is_valid():
            flash("❌ لایسنس منقضی شده است", "error")
            return redirect('/license')
    except:
        flash("❌ خطا در بررسی لایسنس", "error")
        return redirect('/license')
    
    return None
# در بالای routes.py
def update_cash_balance(amount, transaction_type, reference_id=None, description=""):
    """به‌روزرسانی موجودی نقدی و ثبت تراکنش"""
    from app.models import CashBalance, CashTransaction
    from flask_login import current_user
    from datetime import datetime
    
    try:
        cash = CashBalance.query.first()
        if not cash:
            cash = CashBalance(amount=0, updated_by=current_user.id)
            db.session.add(cash)
            db.session.flush()
        
        old_balance = cash.amount
        cash.amount += amount
        cash.updated_by = current_user.id
        cash.last_updated = datetime.now()
        
        # ثبت تراکنش
        transaction = CashTransaction(
            amount=amount,
            transaction_type=transaction_type,
            description=description,
            reference_id=reference_id,
            balance_before=old_balance,
            balance_after=cash.amount,
            created_by=current_user.id
        )
        db.session.add(transaction)
        
        type_persian = {
            'sale': 'فروش',
            'expense': 'مصرف',
            'withdrawal': 'برداشت',
            'adjustment': 'تنظیم',
            'initial': 'موجودی اولیه',
            'return': 'مرجوعی'  # ✅ اضافه شد
        }.get(transaction_type, transaction_type)
        
        print(f"✅ {type_persian}: {amount:+,.0f} افغانی (موجودی: {old_balance:,.0f} → {cash.amount:,.0f})")
        return True
        
    except Exception as e:
        print(f"❌ خطا در به‌روزرسانی موجودی نقدی: {e}")
        import traceback
        traceback.print_exc()
        return False
# در بالای routes.py، بعد از importها و قبل از سایر توابع
# قفل کل سیستم بدون لایسنس - اصلاح شده
# ----------------------------
@main_bp.before_app_request
def block_without_license():
    # اگر کاربر لاگین نیست، بگذار به صفحه لاگین برود
    if not current_user.is_authenticated:
        return
    
    # endpointهای مجاز که نیاز به لایسنس ندارند
    allowed_endpoints = {
        "main_bp.license_page",
        "main_bp.activate_license",
        "main_bp.create_trial_license",
        "main_bp.logout",
        "static"
    }

    # اگر در صفحه مجاز هستیم، اجازه بده
    if request.endpoint in allowed_endpoints:
        return

    # بررسی لایسنس
    license = License.query.first()
    
    # اگر لایسنس وجود ندارد یا معتبر نیست
    if not license or not license.is_valid():
        return redirect(url_for("main_bp.license_page"))
    
    # اگر لایسنس معتبر است، هیچ کاری نکن (به صفحه اصلی برو)

# ----------------------------
# صفحه لایسنس - اصلاح شده# app/routes.py - جایگزینی بخش‌های لایسنس
# app/routes.py
# app/routes.py

@main_bp.route("/license", methods=["GET"])
def license_page():
    """صفحه لایسنس"""
    from app.models import License
    from app.license_config import get_hardware_id
    
    license = License.query.first()
    
    # اگر لایسنس معتبر است و کاربر لاگین است
    if license and license.is_valid() and current_user.is_authenticated:
        flash("✅ لایسنس فعال است", "success")
        return redirect(url_for('main_bp.index'))
    
    # اگر لایسنس معتبر نیست یا کاربر لاگین نیست
    expired = False
    remaining_days = 0
    
    if license:
        try:
            expired = not license.is_valid()
            remaining_days = license.remaining_days()
        except:
            expired = True
    
    return render_template(
        "license.html",
        expired=expired,
        remaining_days=remaining_days,
        hw_id=get_hardware_id()
    )
# app/routes.py

@main_bp.route("/license/activate", methods=["POST"])
def activate_license():
    """فعال‌سازی لایسنس با کلید دریافتی"""
    from app.models import License
    from app.license_config import get_hardware_id, generate_license, get_real_date
    from datetime import timedelta
    from app import db
    
    hardware_id = get_hardware_id()  # ✅ استفاده مستقیم از تابع config
    license_key = request.form.get('license', '').strip().replace('-', '').replace(' ', '')
    
    # تولید کلید معتبر
    valid_key = generate_license(hardware_id)  # ✅ استفاده مستقیم از تابع config
    
    if license_key == valid_key:
        # حذف لایسنس قبلی
        License.query.delete()
        
        real_date = get_real_date()
        
        # ایجاد لایسنس جدید
        new_license = License(
            hw_id=hardware_id,
            expire_at=real_date + timedelta(days=365),
            license_type='full',
            last_valid_date=real_date
        )
        db.session.add(new_license)
        db.session.commit()
        
        flash("✅ لایسنس با موفقیت فعال شد", "success")
        return redirect(url_for('main_bp.index'))
    else:
        flash("❌ کد لایسنس نامعتبر است", "error")
        return redirect(url_for('main_bp.license_page'))
  # app/routes.py

@main_bp.route("/license/trial", methods=["GET"])
@login_required
def create_trial_license():
    """ایجاد لایسنس آزمایشی 30 روزه (فقط یک بار)"""
    from app.models import License
    from app.license_config import get_hardware_id, get_real_date, MAX_TRIAL_COUNT
    from datetime import timedelta
    from app import db
    
    # فقط مدیر
    if current_user.role != 'manager':
        flash("❌ فقط مدیر می‌تواند لایسنس آزمایشی ایجاد کند", "error")
        return redirect(url_for('main_bp.license_page'))
    
    hardware_id = get_hardware_id()
    existing_license = License.query.first()
    
    # بررسی لایسنس قبلی
    if existing_license:
        # اگر لایسنس کامل است
        if existing_license.license_type == 'full':
            flash("❌ شما قبلاً لایسنس کامل را فعال کرده‌اید", "error")
            return redirect(url_for('main_bp.index'))
        
        # اگر تعداد آزمایشی به حد مجاز رسیده
        if existing_license.trial_count >= MAX_TRIAL_COUNT:
            flash("❌ شما قبلاً از لایسنس آزمایشی استفاده کرده‌اید. لطفاً لایسنس کامل خریداری کنید.", "error")
            return redirect(url_for('main_bp.license_page'))
        
        # تمدید لایسنس آزمایشی
        real_date = get_real_date()
        existing_license.expire_at = real_date + timedelta(days=30)
        existing_license.trial_count += 1
        existing_license.last_trial_date = real_date
        existing_license.last_valid_date = real_date
        db.session.commit()
        
        flash(f"✅ لایسنس آزمایشی 30 روزه تمدید شد (دفعه {existing_license.trial_count})", "success")
    else:
        # ایجاد لایسنس آزمایشی جدید
        real_date = get_real_date()
        new_license = License(
            hw_id=hardware_id,
            expire_at=real_date + timedelta(days=30),
            trial_count=1,
            last_trial_date=real_date,
            license_type='trial',
            last_valid_date=real_date
        )
        db.session.add(new_license)
        db.session.commit()
        flash("✅ لایسنس آزمایشی 30 روزه فعال شد", "success")
    
    return redirect(url_for('main_bp.index'))

@main_bp.route("/")
@login_required
def index():
    # دریافت نام عکس ذخیره شده از session
    dashboard_image = session.get('dashboard_image')
    return render_template("index.html", dashboard_image=dashboard_image)


@main_bp.route("/login", methods=["GET", "POST"])
@csrf.exempt   # ⭐ خیلی مهم
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash("ورود موفق بود ✅", "success")

            next_page = request.args.get("next")
            return redirect(next_page or url_for("main_bp.index"))

        flash("نام کاربری یا رمز عبور اشتباه است ❌", "danger")

    return render_template("login.html")


@main_bp.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("main_bp.login"))

# ==================== محصولات ====================
@main_bp.route("/products")
@login_required
def list_products():
    q = request.args.get("q", "").strip()
    batch = request.args.get("batch", "").strip()

    query = Product.query

    if q:
        query = query.filter(Product.name.ilike(f"%{q}%"))

    if batch:
        query = query.filter(Product.batch_no.ilike(f"%{batch}%"))

    products = query.order_by(Product.name).all()
    delete_forms = {p.id: DeleteForm() for p in products}
    today = datetime.today().date()

    return render_template(
        "products.html",
        products=products,
        delete_forms=delete_forms,
        today=today 
    )
 # -----------------------------
@main_bp.route("/products/add", methods=["GET", "POST"])
@login_required
def add_product():
    form = ProductForm()

    CURRENCY = "افغانی"

    try:
        creditors = Creditor.query.order_by(Creditor.name).all()

        form.creditor_id.choices = [
            (0, "انتخاب طلبکار")
        ] + [
            (
                c.id,
                f"{c.name} (قرض داری: {c.current_debt:,.0f} {CURRENCY})"
            )
            for c in creditors
        ]

    except Exception as e:
        flash(
            f"خطا در بارگذاری طلبکاران: {str(e)}",
            "danger"
        )

        form.creditor_id.choices = [
            (0, "انتخاب طلبکار")
        ]

    if form.validate_on_submit():

        try:

            creditor_id = form.creditor_id.data
            if creditor_id == 0:
                creditor_id = None

            # ======================
            # دریافت اطلاعات فرم
            # ======================

            barcode = request.form.get(
                "barcode",
                ""
            ).strip() or None

            batch_no = (
                form.batch_no.data.strip()
                if form.batch_no.data
                else None
            )

            # ✅ تعداد کارتن (اعشاری)
            carton_quantity = float(
                request.form.get(
                    "carton_quantity",
                    0
                ) or 0
            )

            # ✅ تعداد در کارتن
            items_per_carton = float(
                request.form.get(
                    "items_per_carton",
                    1
                ) or 1
            )

            # ✅✅✅ تعداد تکی (جدید) ✅✅✅
            single_quantity = float(
                request.form.get(
                    "single_quantity",
                    0
                ) or 0
            )

            # اعتبارسنجی
            if carton_quantity <= 0 and single_quantity <= 0:
                flash(
                    "❌ حداقل یکی از فیلدهای (تعداد کارتن) یا (تعداد تکی) باید بیشتر از صفر باشد",
                    "danger"
                )
                return render_template(
                    "add_product.html",
                    form=form
                )

            if items_per_carton <= 0 and carton_quantity > 0:
                flash(
                    "❌ تعداد داخل کارتن باید بیشتر از صفر باشد",
                    "danger"
                )
                return render_template(
                    "add_product.html",
                    form=form
                )

            # ======================
            # بررسی بارکد تکراری
            # ======================

            if barcode:

                existing_barcode = Product.query.filter_by(
                    barcode=barcode
                ).first()

                if existing_barcode:

                    flash(
                        f"❌ بارکد قبلاً برای محصول {existing_barcode.name} ثبت شده",
                        "danger"
                    )

                    return render_template(
                        "add_product.html",
                        form=form
                    )

            # ======================
            # محاسبه مجموع تعداد کل
            # فرمول: (تعداد کارتن × تعداد در کارتن) + تعداد تکی
            # ======================

            total_quantity = (
                carton_quantity * items_per_carton
            ) + single_quantity

            print(f"""
                نام: {form.name.data}
                تعداد کارتن: {carton_quantity}
                تعداد در کارتن: {items_per_carton}
                تعداد تکی: {single_quantity}
                تعداد کل: {total_quantity}
            """)

            # ======================
            # ایجاد محصول
            # ======================

            product = Product(
                name=form.name.data,
                batch_no=batch_no,
                barcode=barcode,

                quantity=carton_quantity,                    # تعداد کارتن
                items_per_carton=items_per_carton,            # تعداد در کارتن
                single_quantity=single_quantity,              # ✅✅✅ تعداد تکی (جدید)

                buying_price=form.buying_price.data,
                selling_price=form.selling_price.data,

                unit=form.unit.data,
                category=form.category.data,

                expiry_date=form.expiry_date.data,

                purchase_type=form.purchase_type.data,

                creditor_id=creditor_id,

                purchase_description=form.purchase_description.data,

                is_foreign=False
            )

            db.session.add(product)

            db.session.flush()

            # ======================
            # خرید قرضی
            # ======================

            if (
                form.purchase_type.data == "credit"
                and creditor_id
            ):

                creditor = Creditor.query.get(
                    creditor_id
                )

                if creditor:

                    total_cost = (
                        (form.buying_price.data or 0)
                        * total_quantity
                    )

                    creditor.current_debt += total_cost

                    transaction = DebtTransaction(

                        creditor_id=creditor.id,

                        user_id=current_user.id,

                        amount=total_cost,

                        transaction_type="debt",

                        receipt_number=f"PROD-{product.id}",

                        description=(
                            f"""
خرید محصول: {product.name}

تعداد کارتن: {carton_quantity}
تعداد در کارتن: {items_per_carton}
تعداد تکی: {single_quantity}

مجموع دانه:
{int(total_quantity)}

قیمت خرید:
{product.buying_price:,.0f}
{CURRENCY}
"""
                        ),

                        date_created=datetime.now()
                    )

                    db.session.add(
                        transaction
                    )

            db.session.commit()

            flash(
                f"""
✅ محصول {product.name}
با موفقیت ثبت شد

📦 تعداد کارتن:
{carton_quantity}

📦 دانه در کارتن:
{items_per_carton}

🔸 تعداد تکی:
{single_quantity}

📊 مجموع کل:
{int(total_quantity)}
""",
                "success"
            )

            return redirect(
                url_for(
                    "main_bp.list_products"
                )
            )

        except Exception as e:

            db.session.rollback()

            import traceback
            traceback.print_exc()

            flash(
                f"❌ خطا در ثبت محصول: {str(e)}",
                "danger"
            )

    return render_template(
        "add_product.html",
        form=form
    )
# -----------------------------
@main_bp.route("/customers/advanced")
@login_required
def advanced_customers_page():
    """نمایش صفحه جستجوی پیشرفته مشتریان"""
    return render_template("customer_search.html")  # اسم فایل تمپلیتی که نوشتید

@main_bp.route("/customers/search")
@login_required
def advanced_search_customers():
    name = request.args.get("name", "").strip()
    phone = request.args.get("phone", "").strip()
    debt_status = request.args.get("debt_status", "all")
    page = request.args.get("page", 1, type=int)
    per_page = 10
    
    query = Customer.query
    
    if name:
        query = query.filter(Customer.name.ilike(f"%{name}%"))
    if phone:
        query = query.filter(Customer.phone.ilike(f"%{phone}%"))
    if debt_status == "debtor":
        query = query.filter(Customer.total_debt > 0)
    elif debt_status == "no_debt":
        query = query.filter(Customer.total_debt == 0)
    
    pagination = query.order_by(Customer.name).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    customers_data = [{
        "id": c.id,
        "name": c.name,
        "phone": c.phone or "-",
        "total_debt": float(c.total_debt),
        "created_at": c.created_at.isoformat() if c.created_at else None
    } for c in pagination.items]
    
    return jsonify({
        "customers": customers_data,
        "total": pagination.total,
        "page": page,
        "pages": pagination.pages,
        "has_next": pagination.has_next,
        "has_prev": pagination.has_prev,
        "next_num": pagination.next_num,
        "prev_num": pagination.prev_num
    })
@main_bp.route("/customers")
@login_required
def list_customers():
    search = request.args.get("search", "").strip()
    query = Customer.query

    if search:
        # جستجو بر اساس نام یا شماره تماس
        query = query.filter(
            (Customer.name.ilike(f"%{search}%")) |
            (Customer.phone.ilike(f"%{search}%"))
        )

    customers = query.order_by(Customer.name).all()
    return render_template("customers.html", customers=customers, search=search)

# -----------------------------
# افزودن مشتری
# -----------------------------

@main_bp.route("/customers/add", methods=["GET", "POST"])
@login_required
def add_customer():
    form = CustomerForm()
    if form.validate_on_submit():
        customer = Customer(
            name=form.name.data,
            phone=form.phone.data,
            address=form.address.data,
            total_debt=form.total_debt.data
        )
        db.session.add(customer)
        db.session.commit()
        flash("مشتری جدید ثبت شد ✅", "success")
        return redirect(url_for("main_bp.list_customers"))

    return render_template("add_customer.html", form=form)


# -----------------------------
# ویرایش مشتری
# -----------------------------
@main_bp.route("/customers/edit/<int:customer_id>", methods=["GET", "POST"])
@login_required
def edit_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    form = CustomerForm(obj=customer)
    if form.validate_on_submit():
        form.populate_obj(customer)
        db.session.commit()
        flash("مشتری ویرایش شد ✅", "success")
        return redirect(url_for("main_bp.list_customers"))

    return render_template("edit_customer.html", form=form, customer=customer)


# -----------------------------
# حذف مشتری
# -----------------------------
@main_bp.route("/customers/delete/<int:customer_id>", methods=["POST"])
@login_required
def delete_customer(customer_id):
    try:
        customer = Customer.query.get_or_404(customer_id)

        # جلوگیری از حذف مشتری با سابقه فروش
        if customer.sales:
            flash("❌ امکان حذف مشتری با سابقه فروش وجود ندارد", "error")
            return redirect(url_for("main_bp.list_customers"))

        db.session.delete(customer)
        db.session.commit()
        flash("✅ مشتری با موفقیت حذف شد", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"❌ خطا در حذف مشتری: {str(e)}", "error")

    return redirect(url_for("main_bp.list_customers"))

@main_bp.route("/sales/quick_add_customer", methods=["POST"])
@login_required
def quick_add_customer():
    """افزودن سریع مشتری جدید از صفحه فروش"""
    try:
        data = request.get_json()
        
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'message': 'نام مشتری الزامی است'})
        
        # ✅ ایجاد مشتری جدید در دیتابیس
        new_customer = Customer(
            name=name,
            phone=data.get('phone', '').strip() or None,
            address=data.get('address', '').strip() or None,
            total_debt=0.0  # مشتری جدید قرض داری ندارد
        )
        
        db.session.add(new_customer)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'✅ مشتری {name} با موفقیت اضافه شد',
            'customer': {
                'id': new_customer.id,
                'name': new_customer.name,
                'phone': new_customer.phone or ''
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})
# -----------------------------
# جزئیات مشتری
# -----------------------------
@main_bp.route("/customers/<int:customer_id>")
@login_required
def customer_detail(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    sales = Sale.query.filter_by(customer_id=customer.id).order_by(Sale.date.desc()).all()
    payments = DebtPayment.query.filter_by(customer_id=customer.id).order_by(DebtPayment.date.desc()).all()
    return render_template("customer_detail.html", customer=customer, sales=sales, payments=payments)


# -----------------------------
# پرداخت قرض داری مشتری
# -----------------------------
@main_bp.route("/customers/pay_debt/<int:customer_id>", methods=["GET", "POST"])
@login_required
def pay_debt(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    form = DebtPaymentForm()

    # دریافت موجودی فعلی صندوق
    cash = CashBalance.query.first()
    if not cash:
        cash = CashBalance(amount=0)
        db.session.add(cash)
        db.session.commit()

    if request.method == "POST":
        if form.validate_on_submit():
            amount = float(form.amount.data)
            
            if amount <= 0:
                flash("❌ مقدار پرداخت باید مثبت باشد", "danger")
            elif amount > customer.total_debt:
                flash(f"❌ مبلغ پرداختی ({amount:,.0f}) بیشتر از قرض فعلی ({customer.total_debt:,.0f}) است", "danger")
            else:
                # ========== 1. ثبت پرداخت ==========
                payment = DebtPayment(
                    customer_id=customer.id,
                    amount=amount,
                    receipt_number=form.receipt_number.data,
                    date=datetime.utcnow(),
                    created_by=current_user.id
                )
                db.session.add(payment)
                
                # ========== 2. کاهش بدهی مشتری ==========
                customer.total_debt -= amount
                
                # ========== 3. افزایش موجودی نقدی صندوق ✅ ==========
                old_cash_balance = cash.amount
                cash.amount += amount
                
                # ثبت تراکنش نقدی
                cash_transaction = CashTransaction(
                    amount=amount,
                    transaction_type='debt_payment',
                    description=f"پرداخت قرض از مشتری {customer.name} - مبلغ: {amount:,.0f} افغانی - رسید: {form.receipt_number.data or 'بدون رسید'}",
                    balance_before=old_cash_balance,
                    balance_after=cash.amount,
                    created_by=current_user.id,
                    reference_id=payment.id
                )
                db.session.add(cash_transaction)
                
                db.session.commit()
                
                flash(f"✅ پرداخت به مبلغ {amount:,.0f} افغانی از {customer.name} ثبت شد", "success")
                flash(f"💰 موجودی فعلی صندوق: {cash.amount:,.0f} افغانی", "info")
                
                return redirect(url_for("main_bp.customer_detail", customer_id=customer.id))

    return render_template("pay_debt.html", form=form, customer=customer, cash_balance=cash.amount)
# -----------------------------
# لیست بدهکاران
# -----------------------------
@main_bp.route("/customers/debtors")
@login_required
def list_debtors():
    debtors = Customer.query.filter(Customer.total_debt > 0).order_by(Customer.total_debt.desc()).all()
    return render_template("debtors.html", debtors=debtors)

@main_bp.route("/sales/add", methods=["GET", "POST"])
@login_required
def add_sale():
    from datetime import datetime
    from app import db
    from app.models import Sale, SaleItem, Customer, Product, CashBalance

    # =========================
    # GET - نمایش فرم
    # =========================
    if request.method == "GET":
        customers = Customer.query.order_by(Customer.name).all()
        return render_template(
            "add_sale.html",
            customers=customers,
            title="ثبت فروش جدید"
        )

    # =========================
    # POST - ثبت فروش
    # =========================
    try:
        # دریافت اطلاعات از فرم
        customer_id = request.form.get("customer_id", type=int)
        if customer_id == 0:
            customer_id = None

        product_ids = request.form.getlist("product_id[]")
        carton_quantities = request.form.getlist("carton_quantity[]")
        single_quantities = request.form.getlist("single_quantity[]")
        selling_prices = request.form.getlist("selling_price[]")
        discount_percents = request.form.getlist("discount_percent[]")
        amount_paid = float(request.form.get("amount_paid", 0) or 0)
        total_discount_percent = float(request.form.get("total_discount_percent", 0) or 0)

        # اعتبارسنجی: حداقل یک محصول
        valid_products = [p for p in product_ids if p]
        if not valid_products:
            flash("❌ حداقل یک محصول انتخاب کنید", "danger")
            return redirect(url_for("main_bp.add_sale"))

        # =========================
        # تولید شماره فاکتور
        # =========================
        last_sale = Sale.query.order_by(Sale.id.desc()).first()
        invoice_number = "1"
        if last_sale and last_sale.invoice_number:
            try:
                invoice_number = str(int(last_sale.invoice_number) + 1)
            except:
                invoice_number = str(last_sale.id + 1)

        # =========================
        # ایجاد فروش جدید
        # =========================
        sale = Sale(
            invoice_number=invoice_number,
            customer_id=customer_id,
            total_amount=0,
            total_discount=0,
            final_amount=0,
            amount_paid=amount_paid,
            remaining_debt=0,
            created_by=current_user.id,
            date=datetime.now()
        )
        db.session.add(sale)
        db.session.flush()  # برای گرفتن sale.id

        # متغیرهای محاسباتی
        total_amount = 0
        total_items_discount = 0
        sale_items_list = []

        # =========================
        # حلقه ثبت آیتم‌های فروش
        # =========================
        for i, product_id in enumerate(product_ids):
            if not product_id:
                continue

            product = Product.query.get(int(product_id))
            if not product:
                continue

            # دریافت مقادیر
            carton_qty = int(carton_quantities[i] or 0)
            single_qty = int(single_quantities[i] or 0)
            selling_price = float(selling_prices[i] or 0)
            discount_percent = float(discount_percents[i] or 0)

            # اگر تعداد صفر است، رد کن
            if carton_qty == 0 and single_qty == 0:
                continue

            # اطلاعات محصول
            items_per_carton = float(product.items_per_carton or 1)
            
            # =========================
            # محاسبه موجودی فعلی (به روش مستقیم)
            # =========================
            current_cartons = float(product.quantity or 0)
            current_singles = int(product.single_quantity or 0)
            current_stock = int(current_cartons * items_per_carton + current_singles)

            # محاسبه تعداد فروش به دانه
            sold_items = int(carton_qty * items_per_carton + single_qty)

            # لاگ برای دیباگ
            print(f"\n🔍 بررسی محصول: {product.name}")
            print(f"   کارتن فعلی: {current_cartons}, هر کارتن: {items_per_carton} عدد")
            print(f"   تک فعلی: {current_singles}")
            print(f"   موجودی کل: {current_stock} عدد")
            print(f"   فروش: {carton_qty} کارتن + {single_qty} تک = {sold_items} عدد")

            # =========================
            # بررسی موجودی کافی
            # =========================
            if sold_items > current_stock:
                db.session.rollback()
                flash(
                    f"❌ موجودی {product.name} کافی نیست! "
                    f"(موجودی: {current_stock} عدد، درخواستی: {sold_items} عدد)",
                    "danger"
                )
                return redirect(url_for("main_bp.add_sale"))

            # =========================
            # محاسبه موجودی جدید
            # =========================
            remaining_total = current_stock - sold_items
            new_cartons = int(remaining_total // items_per_carton)
            new_singles = int(remaining_total % items_per_carton)

            # به‌روزرسانی موجودی محصول
            product.quantity = float(new_cartons)
            product.single_quantity = new_singles

            print(f"   ✅ موجودی جدید: {new_cartons} کارتن + {new_singles} تک = {remaining_total} عدد")

            # =========================
            # محاسبات مالی آیتم
            # =========================
            buying_price = float(product.buying_price or 0)
            item_total = sold_items * selling_price
            item_discount = item_total * discount_percent / 100
            item_final = item_total - item_discount
            
            # محاسبه مفاد
            profit_per_item = selling_price - buying_price
            total_profit = profit_per_item * sold_items

            # =========================
            # ایجاد آیتم فروش
            # =========================
            sale_item = SaleItem(
                sale_id=sale.id,
                product_id=product.id,
                quantity=sold_items,
                selling_price=selling_price,
                discount_percent=discount_percent,
                discount_amount=item_discount,
                final_amount=item_final,
                profit=total_profit
            )
            db.session.add(sale_item)
            sale_items_list.append(sale_item)

            # جمع کل
            total_amount += item_total
            total_items_discount += item_discount

        # =========================
        # محاسبات نهایی فاکتور
        # =========================
        # تخفیف کل فاکتور
        invoice_discount = ((total_amount - total_items_discount) * total_discount_percent / 100)
        final_discount = total_items_discount + invoice_discount
        final_amount = total_amount - final_discount
        remaining = final_amount - amount_paid
        if remaining < 0:
            remaining = 0

        # به‌روزرسانی فروش
        sale.total_amount = total_amount
        sale.total_discount = final_discount
        sale.final_amount = final_amount
        sale.remaining_debt = remaining

        # =========================
        # افزایش موجودی نقدی
        # =========================
        if amount_paid > 0:
            cash = CashBalance.query.first()
            if not cash:
                cash = CashBalance(amount=0)
                db.session.add(cash)
            cash.amount += amount_paid

        # =========================
        # به‌روزرسانی قرض داری مشتری
        # =========================
        if customer_id and remaining > 0:
            customer = Customer.query.get(customer_id)
            if customer:
                customer.total_debt += remaining

        # =========================
        # ذخیره نهایی در دیتابیس
        # =========================
        db.session.commit()

        # =========================
        # لاگ موفقیت
        # =========================
        print("\n" + "="*60)
        print(f"✅ فروش با موفقیت ثبت شد | فاکتور شماره: {invoice_number}")
        print(f"   مبلغ کل: {total_amount:,.0f}")
        print(f"   تخفیف کل: {final_discount:,.0f}")
        print(f"   مبلغ نهایی: {final_amount:,.0f}")
        print(f"   پرداختی: {amount_paid:,.0f}")
        print(f"   باقیمانده: {remaining:,.0f}")
        print("\n📦 خلاصه موجودی پس از فروش:")
        for item in sale_items_list:
            if item.product:
                p = item.product
                items_per_carton = float(p.items_per_carton or 1)
                total_stock = int((p.quantity or 0) * items_per_carton + (p.single_quantity or 0))
                print(f"   {p.name}: {p.quantity:.0f} کارتن + {p.single_quantity} تک = {total_stock} عدد")
        print("="*60)

        flash(f"✅ فروش ثبت شد | فاکتور {invoice_number}", "success")
        return redirect(url_for("main_bp.list_sales"))

    except Exception as e:
        db.session.rollback()
        print(f"\n❌ خطا در ثبت فروش: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f"❌ خطا در ثبت فروش: {str(e)}", "danger")
        return redirect(url_for("main_bp.add_sale"))
          
@main_bp.route("/debt/report")
@login_required
def debt_report():
    """گزارش جامع قرض‌ها (دریافتی و پرداختی)"""
    from datetime import datetime, timedelta
    import calendar
    
    filter_type = request.args.get('filter', 'month')
    selected_month = request.args.get('month', type=int)
    
    today = datetime.now().date()
    
    # تعیین بازه زمانی
    if filter_type == 'today':
        start_date = today
        end_date = today
    elif filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif filter_type == 'month':
        if selected_month:
            start_date = datetime(today.year, selected_month, 1).date()
            last_day = calendar.monthrange(today.year, selected_month)[1]
            end_date = datetime(today.year, selected_month, last_day).date()
        else:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = datetime(2020, 1, 1).date()
        end_date = today
    
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    # ===== قرض‌های دریافتی (مشتریان) =====
    received_loans = Sale.query.filter(
        Sale.customer_id.isnot(None),
        Sale.date >= start_datetime,
        Sale.date <= end_datetime,
        Sale.remaining_debt > 0
    ).all()
    
    total_received = sum(s.remaining_debt for s in received_loans)
    received_count = len(received_loans)
    
    # ===== قرض‌های پرداختی (طلبکاران) =====
    paid_loans = DebtTransaction.query.filter(
        DebtTransaction.transaction_type == 'debt',
        DebtTransaction.date_created >= start_datetime,
        DebtTransaction.date_created <= end_datetime
    ).all()
    
    total_paid = sum(t.amount for t in paid_loans)
    paid_count = len(paid_loans)
    
    # ===== پرداخت‌های انجام شده به طلبکاران =====
    payments = DebtTransaction.query.filter(
        DebtTransaction.transaction_type == 'payment',
        DebtTransaction.date_created >= start_datetime,
        DebtTransaction.date_created <= end_datetime
    ).all()
    
    total_payments = sum(t.amount for t in payments)
    payments_count = len(payments)
    
    # ===== مجموع قرض داری‌ها =====
    total_customer_debt = db.session.query(db.func.sum(Customer.total_debt)).scalar() or 0
    total_creditor_debt = db.session.query(db.func.sum(Creditor.current_debt)).scalar() or 0
    
    # ===== لیست ماه‌ها =====
    months_list = []
    month_names = [ 'جدی', 'دلو', 'حوت','حمل','ثور', 'جوزا', 'سرطان', 'اسد', 'سنبله',
                   'میزان', 'عقرب', 'قوس']
    
    for i, name in enumerate(month_names, 1):
        months_list.append((i, f"{i} - {name}"))
    
    return render_template(
        "debt_report.html",
        filter_type=filter_type,
        selected_month=selected_month,
        months_list=months_list,
        start_date=start_datetime,
        end_date=end_datetime,
        
        # قرض‌های دریافتی
        received_loans=received_loans,
        total_received=total_received,
        received_count=received_count,
        
        # قرض‌های پرداختی
        paid_loans=paid_loans,
        total_paid=total_paid,
        paid_count=paid_count,
        
        # پرداخت‌ها
        payments=payments,
        total_payments=total_payments,
        payments_count=payments_count,
        
        # خلاصه
        net_debt=total_paid - total_payments,
        total_creditor_debt=total_creditor_debt,
        total_customer_debt=total_customer_debt
    )
@main_bp.route("/sales")
@login_required
def list_sales():
    filter_type = request.args.get("filter", "all")
    
    # ساخت query پایه
    query = Sale.query.options(
        joinedload(Sale.customer),
        joinedload(Sale.user),
        joinedload(Sale.items).joinedload(SaleItem.product)
    )
    
    # اعمال فیلتر
    if filter_type == "paid":
        query = query.filter(Sale.remaining_debt == 0)
    elif filter_type == "debt":
        query = query.filter(Sale.remaining_debt > 0)
    
    # دریافت همه نتایج
    sales = query.order_by(Sale.date.desc()).all()
    
    return render_template(
        "sales.html", 
        sales=sales,
        selected_filter=filter_type
    )
@main_bp.route("/api/add_foreign_item", methods=["POST"])
@login_required
def add_foreign_item():
    """ساده‌ترین راه برای افزودن کالای خارجی"""
    try:
        # داده‌های ساده
        name = request.json.get('name')
        price = float(request.json.get('price', 0))
        quantity = float(request.json.get('quantity', 1))
        
        if not name or price <= 0:
            return jsonify({'success': False, 'message': 'نام و قیمت الزامی است'}), 400
        
        # ۱. کالای خارجی رو ذخیره کن
        fp = ForeignProduct(
            name=name,
            batch_no=f"FP-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            buying_price=price * 0.8,  # فرضی
            selling_price=price,
            quantity=quantity,
            added_by=current_user.id
        )
        fp.calculate_profit()
        
        db.session.add(fp)
        db.session.flush()
        
        # ۲. محصول موقت برای dropdown
        temp = Product(
            name=f"{name} (خارجی)",
            selling_price=price,
            quantity=quantity,
            unit="عدد"
        )
        
        db.session.add(temp)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'id': temp.id,
            'name': temp.name,
            'price': temp.selling_price
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
# Route view_invoice
from sqlalchemy.orm import joinedload
from app.models import SaleItem 

@main_bp.route("/sales/<int:sale_id>")
@login_required
def view_invoice(sale_id):
    import time
    start = time.time()
    try:
        sale = Sale.query.options(
            joinedload(Sale.items).joinedload(SaleItem.product),
            joinedload(Sale.customer)
        ).filter_by(id=sale_id).first_or_404()
        customer = sale.customer

        # ===== مرحله ۲: محاسبه مقادیر هر آیتم =====
        for item in sale.items:
            item.total_amount = item.quantity * item.selling_price
            item.discount_amount = item.total_amount * (item.discount_percent or 0) / 100
            item.final_amount = item.total_amount - item.discount_amount

        # محاسبه مجموع فاکتور
        sale.total_amount = sum(item.total_amount for item in sale.items)
        sale.total_discount = sum(item.discount_amount for item in sale.items)
        sale.final_total = sum(item.final_amount for item in sale.items)


        # ===== محاسبه قرض داری‌های قبلی =====
        if customer:
            previous_sales = Sale.query.filter(
                Sale.customer_id == customer.id,
                Sale.date < sale.date
            ).all()
            previous_remaining = sum((s.total_amount - s.amount_paid) for s in previous_sales)
            total_remaining = previous_remaining + sale.remaining_debt
        else:
            previous_remaining = 0
            total_remaining = sale.remaining_debt
        print("Execution time:", time.time() - start)
        return render_template(
            "invoice.html",
            sale=sale,
            previous_remaining=previous_remaining,
            total_remaining=total_remaining
        )
    except Exception as e:
        flash(f"خطا در بارگذاری فاکتور: {str(e)}", "danger")
        return redirect(url_for("main_bp.list_sales"))
    

@main_bp.route("/sales/<int:sale_id>/print")
@login_required
def print_invoice(sale_id):
    try:
        sale = Sale.query.get_or_404(sale_id)
        customer = sale.customer

        if customer:
            

                previous_remaining = db.session.query(
                 func.coalesce(func.sum(Sale.total_amount - Sale.amount_paid), 0)
                ).filter(
                 Sale.customer_id == customer.id,
                 Sale.date < sale.date
                ).scalar()
 
                total_remaining = previous_remaining + sale.remaining_debt
        else:
            previous_remaining = 0
            total_remaining = sale.remaining_debt

        return render_template(
            "invoice_print.html",
            sale=sale,
            previous_remaining=previous_remaining,
            total_remaining=total_remaining
        )
    except Exception as e:
        flash(f"خطا در پرینت فاکتور: {str(e)}", "danger")
        return redirect(url_for("main_bp.list_sales"))
@main_bp.route("/sales/<int:sale_id>/edit_full", methods=["GET", "POST"])
@login_required
def edit_sale_full(sale_id):
    from app.models import Sale, SaleItem, Customer, Product, CashBalance, CashTransaction
    from app.forms import EditSaleFullForm
    from datetime import datetime
    
    sale = Sale.query.get_or_404(sale_id)
    
    # دریافت موجودی صندوق
    cash = CashBalance.query.first()
    if not cash:
        cash = CashBalance(amount=0)
        db.session.add(cash)
        db.session.commit()
    
    if request.method == "POST":
        try:
            print("="*80)
            print(f"🟢 شروع ویرایش فاکتور {sale.invoice_number} (ID: {sale.id})")
            print("="*80)
            
            # ============================================================
            # مرحله 1: ذخیره اطلاعات قبلی
            # ============================================================
            old_customer_id = sale.customer_id
            old_remaining_debt = sale.remaining_debt
            old_final_amount = sale.final_amount
            old_amount_paid = sale.amount_paid
            old_items_list = list(sale.items)
            
            print(f"\n📋 اطلاعات قبلی:")
            print(f"   مبلغ نهایی قبلی: {old_final_amount:,.0f}")
            print(f"   پرداختی قبلی: {old_amount_paid:,.0f}")
            print(f"   بدهی قبلی: {old_remaining_debt:,.0f}")
            
            # ✅ دریافت مقادیر جدید از فرم
            customer_id = request.form.get("customer_id")
            customer_id = int(customer_id) if customer_id and customer_id != "0" else None
            
            total_discount_percent = float(request.form.get("total_discount_percent", 0) or 0)
            amount_paid = float(request.form.get("amount_paid", 0) or 0)  # ✅ مبلغ پرداختی جدید
            description = request.form.get("description", "")
            
            print(f"\n📋 اطلاعات جدید:")
            print(f"   تخفیف کل: {total_discount_percent}%")
            print(f"   پرداختی جدید: {amount_paid:,.0f}")
            
            # ============================================================
            # مرحله 2: برگرداندن موجودی قبلی به انبار
            # ============================================================
            print(f"\n🔄 مرحله 1: برگرداندن موجودی قبلی به انبار")
            
            for item in old_items_list:
                product = Product.query.get(item.product_id)
                if product:
                    items_per_carton = product.items_per_carton or 1
                    total_sold_items = int(item.quantity)
                    
                    cartons_to_return = total_sold_items // items_per_carton
                    singles_to_return = total_sold_items % items_per_carton
                    
                    product.quantity = (product.quantity or 0) + cartons_to_return
                    product.single_quantity = (product.single_quantity or 0) + singles_to_return
                    
                    print(f"   ↩️ {product.name}: +{cartons_to_return} کارتن + {singles_to_return} تک")
            
            # ============================================================
            # مرحله 3: حذف آیتم‌های قبلی
            # ============================================================
            for item in old_items_list:
                db.session.delete(item)
            
            # ============================================================
            # مرحله 4: دریافت داده‌های جدید از فرم
            # ============================================================
            product_ids = request.form.getlist("product_id[]")
            carton_quantities = request.form.getlist("carton_quantity[]")
            single_quantities = request.form.getlist("single_quantity[]")
            prices = request.form.getlist("price[]")
            discounts = request.form.getlist("discount[]")
            
            total_amount = 0
            total_item_discount = 0
            
            print(f"\n📦 مرحله 2: ثبت فروش جدید")
            
            # ============================================================
            # مرحله 5: پردازش محصولات جدید
            # ============================================================
            for i, product_id in enumerate(product_ids):
                if not product_id:
                    continue
                
                product = Product.query.get(int(product_id))
                if not product:
                    continue
                
                carton_qty = float(carton_quantities[i]) if i < len(carton_quantities) and carton_quantities[i] else 0
                single_qty = float(single_quantities[i]) if i < len(single_quantities) and single_quantities[i] else 0
                units_per_carton = product.items_per_carton or 1
                
                total_quantity = int((carton_qty * units_per_carton) + single_qty)
                
                if total_quantity <= 0:
                    continue
                
                price = float(prices[i]) if i < len(prices) and prices[i] else product.selling_price
                discount = float(discounts[i]) if i < len(discounts) and discounts[i] else 0
                
                # بررسی موجودی کافی
                current_cartons = product.quantity or 0
                current_singles = product.single_quantity or 0
                current_stock = int((current_cartons * units_per_carton) + current_singles)
                
                if current_stock < total_quantity:
                    db.session.rollback()
                    flash(f"❌ موجودی محصول {product.name} کافی نیست!", "danger")
                    return redirect(url_for("main_bp.edit_sale_full", sale_id=sale.id))
                
                # کاهش موجودی
                new_total_stock = current_stock - total_quantity
                new_cartons = new_total_stock // units_per_carton
                new_singles = new_total_stock % units_per_carton
                
                product.quantity = float(new_cartons)
                product.single_quantity = int(new_singles)
                
                # محاسبات مالی
                item_total = total_quantity * price
                item_discount = item_total * (discount / 100)
                final_total = item_total - item_discount
                
                buying_price = product.buying_price or 0
                total_profit = (price - buying_price) * total_quantity
                
                # ایجاد آیتم جدید
                sale_item = SaleItem(
                    sale_id=sale.id,
                    product_id=product.id,
                    quantity=total_quantity,
                    selling_price=price,
                    discount_percent=discount,
                    discount_amount=item_discount,
                    final_amount=final_total,
                    profit=total_profit
                )
                db.session.add(sale_item)
                
                total_amount += item_total
                total_item_discount += item_discount
                
                print(f"   ✅ {product.name}: {total_quantity} عدد × {price:,.0f} = {item_total:,.0f}")
            
            # ============================================================
            # مرحله 6: محاسبه مبالغ نهایی
            # ============================================================
            total_discount_amount = total_amount * (total_discount_percent / 100)
            final_amount = total_amount - total_item_discount - total_discount_amount
            
            # محاسبه بدهی جدید
            new_remaining_debt = final_amount - amount_paid
            is_overpaid = new_remaining_debt < 0
            
            if is_overpaid:
                refund_amount = abs(new_remaining_debt)
                new_remaining_debt = 0
                # اضافه پرداخت به صندوق برگردانده نمی‌شود، بلکه به عنوان بستانکاری مشتری ثبت می‌شود
                print(f"   ⚠️ اضافه پرداخت: {refund_amount:,.0f} افغانی")
            else:
                refund_amount = 0
            
            # ============================================================
            # مرحله 7: ✅ به‌روزرسانی موجودی صندوق (مهم!)
            # ============================================================
            
            # محاسبه تغییر در موجودی صندوق
            # موجودی صندوق باید بر اساس مبلغ پرداختی جدید به‌روز شود
            cash_change = amount_paid - old_amount_paid
            
            if cash_change != 0:
                old_cash = cash.amount
                cash.amount += cash_change
                
                # ثبت تراکنش مالی
                if cash_change > 0:
                    trans_type = 'sale_payment'
                    trans_desc = f"پرداخت اضافه برای ویرایش فاکتور {sale.invoice_number}"
                else:
                    trans_type = 'sale_refund'
                    trans_desc = f"برگشت وجه از ویرایش فاکتور {sale.invoice_number}"
                
                cash_transaction = CashTransaction(
                    amount=cash_change,
                    transaction_type=trans_type,
                    description=trans_desc,
                    balance_before=old_cash,
                    balance_after=cash.amount,
                    created_by=current_user.id,
                    reference_id=sale.id
                )
                db.session.add(cash_transaction)
                
                print(f"\n💰 تغییرات صندوق:")
                print(f"   پرداختی قبلی: {old_amount_paid:,.0f}")
                print(f"   پرداختی جدید: {amount_paid:,.0f}")
                print(f"   تغییر: {cash_change:+,.0f}")
                print(f"   موجودی قبل: {old_cash:,.0f}")
                print(f"   موجودی بعد: {cash.amount:,.0f}")
            
            # ============================================================
            # مرحله 8: به‌روزرسانی فیلدهای فاکتور
            # ============================================================
            sale.customer_id = customer_id
            sale.total_amount = total_amount
            sale.total_discount = total_item_discount + total_discount_amount
            sale.final_amount = final_amount
            sale.amount_paid = amount_paid  # ✅ مبلغ پرداختی جدید
            sale.remaining_debt = new_remaining_debt
            sale.description = description
            
            if hasattr(sale, 'total_discount_percent'):
                sale.total_discount_percent = total_discount_percent
            
            # به‌روزرسانی تاریخ
            if request.form.get("date"):
                try:
                    sale.date = datetime.strptime(request.form.get("date"), "%Y-%m-%d")
                except:
                    pass
            
            # ============================================================
            # مرحله 9: اصلاح بدهی مشتری
            # ============================================================
            # حذف قرض از مشتری قبلی
            if old_customer_id:
                old_customer = Customer.query.get(old_customer_id)
                if old_customer:
                    old_customer.total_debt = max(0, (old_customer.total_debt or 0) - old_remaining_debt)
                    print(f"   👤 مشتری قبلی: بدهی جدید = {old_customer.total_debt:,.0f}")
            
            # اضافه کردن قرض به مشتری جدید
            if sale.customer_id and sale.remaining_debt > 0:
                new_customer = Customer.query.get(sale.customer_id)
                if new_customer:
                    new_customer.total_debt = (new_customer.total_debt or 0) + sale.remaining_debt
                    print(f"   👤 مشتری جدید: بدهی جدید = {new_customer.total_debt:,.0f}")
            
            # ============================================================
            # مرحله 10: ذخیره نهایی
            # ============================================================
            db.session.commit()
            
            print("\n" + "="*80)
            print(f"✅ فاکتور {sale.invoice_number} با موفقیت ویرایش شد")
            print(f"💰 موجودی نهایی صندوق: {cash.amount:,.0f} افغانی")
            print(f"💳 مبلغ پرداختی: {amount_paid:,.0f} افغانی")
            print(f"📝 بدهی باقیمانده: {new_remaining_debt:,.0f} افغانی")
            print("="*80)
            
            flash("✅ فاکتور با موفقیت ویرایش شد", "success")
            if cash_change > 0:
                flash(f"💰 {cash_change:+,.0f} افغانی به صندوق اضافه شد", "success")
            elif cash_change < 0:
                flash(f"💰 {abs(cash_change):,.0f} افغانی از صندوق کسر شد", "warning")
            flash(f"💰 موجودی فعلی صندوق: {cash.amount:,.0f} افغانی", "info")
            
            return redirect(url_for("main_bp.view_invoice", sale_id=sale.id))
            
        except Exception as e:
            db.session.rollback()
            print(f"\n❌ خطا: {str(e)}")
            import traceback
            traceback.print_exc()
            flash(f"❌ خطا در ویرایش فاکتور: {str(e)}", "danger")
            return redirect(url_for("main_bp.edit_sale_full", sale_id=sale.id))
    
    # ============================================================
    # GET REQUEST
    # ============================================================
    form = EditSaleFullForm()
    form.customer_id.choices = [(0, "بدون مشتری")] + [(c.id, f"{c.name} - {c.phone or 'بدون شماره'}") for c in Customer.query.all()]
    form.customer_id.data = sale.customer_id if sale.customer_id else 0
    
    products = Product.query.all()
    
    return render_template(
        "edit_sale_full.html", 
        sale=sale, 
        form=form, 
        products=products, 
        cash_balance=cash.amount
    )
@main_bp.route("/sales/search", methods=["GET", "POST"])
@login_required
def sales_search():
    sales = []
    total_amount = 0
    total_paid = 0
    total_debt = 0
    
    # دریافت پارامترهای جستجو از URL
    invoice_number = request.args.get('invoice_number', '').strip()
    customer_name = request.args.get('customer_name', '').strip()
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    print(f"🔍 پارامترهای جستجو:")
    print(f"   شماره فاکتور: {invoice_number}")
    print(f"   نام مشتری: {customer_name}")
    print(f"   از تاریخ: {start_date}")
    print(f"   تا تاریخ: {end_date}")
    
    # فقط اگر پارامتر جستجو وجود دارد، جستجو کنیم
    if invoice_number or customer_name or start_date or end_date:
        query = Sale.query
        
        if invoice_number:
            query = query.filter(Sale.invoice_number.ilike(f"%{invoice_number}%"))
            print(f"✅ فیلتر شماره فاکتور اعمال شد: {invoice_number}")
        
        if customer_name:
            query = query.join(Customer).filter(Customer.name.ilike(f"%{customer_name}%"))
            print(f"✅ فیلتر نام مشتری اعمال شد: {customer_name}")
        
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(Sale.date >= start_datetime)
                print(f"✅ فیلتر از تاریخ اعمال شد: {start_date}")
            except ValueError:
                print(f"❌ خطا در تاریخ شروع: {start_date}")
        
        if end_date:
            try:
                # اضافه کردن یک روز به end_date برای شامل شدن آن روز
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(Sale.date < end_datetime)
                print(f"✅ فیلتر تا تاریخ اعمال شد: {end_date}")
            except ValueError:
                print(f"❌ خطا در تاریخ پایان: {end_date}")
        
        sales = query.order_by(Sale.date.desc()).all()
        print(f"📊 تعداد نتایج پیدا شده: {len(sales)}")
        
        # محاسبات آماری
        total_amount = sum(sale.total_amount for sale in sales)
        total_paid = sum(sale.amount_paid for sale in sales)
        total_debt = sum(sale.remaining_debt for sale in sales)
    
    # ایجاد فرم (برای نمایش اولیه)
    form = SaleSearchForm()
    
    return render_template("sales_search.html", 
                         form=form, 
                         sales=sales,
                         total_amount=total_amount,
                         total_paid=total_paid,
                         total_debt=total_debt)
# ==================== API کالای خارجی ====================

@main_bp.route("/api/foreign_product", methods=["POST"])
@login_required
def add_foreign_product_api():
    """API برای افزودن کالای خارجی"""
    try:
        data = request.get_json()
        
        # اعتبارسنجی
        if not data.get('name') or not data.get('selling_price'):
            return jsonify({'success': False, 'message': 'نام و قیمت الزامی است'}), 400
        
        # ایجاد کالای خارجی (بدون quantity)
        foreign_product = ForeignProduct(
            name=data['name'],
            batch_no=data.get('batch_no', f'FP-{datetime.now().strftime("%Y%m%d%H%M%S")}'),
            buying_price=float(data.get('buying_price', float(data['selling_price']) * 0.8)),
            selling_price=float(data['selling_price']),
            unit=data.get('unit', 'عدد'),
            description=data.get('description', ''),
            added_by=current_user.id
        )
        foreign_product.calculate_profit()
        
        db.session.add(foreign_product)
        db.session.flush()
        
        # ایجاد محصول برای dropdown (اینجا quantity داریم)
        temp_product = Product(
            name=f"{data['name']} (خارجی)",
            batch_no=foreign_product.batch_no,
            buying_price=foreign_product.buying_price,
            selling_price=foreign_product.selling_price,
            quantity=float(data.get('quantity', 1)),
            unit=foreign_product.unit,
            purchase_description="کالای خارجی"
        )
        
        db.session.add(temp_product)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'کالای خارجی ثبت شد',
            'data': {
                'temp_id': temp_product.id,
                'foreign_id': foreign_product.id,
                'name': temp_product.name,
                'selling_price': temp_product.selling_price,
                'quantity': temp_product.quantity
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
        
@main_bp.route("/api/foreign_products/list")
@login_required
def list_foreign_products():
    """دریافت لیست کالاهای خارجی موجود"""
    try:
        foreign_products = ForeignProduct.query.order_by(ForeignProduct.created_at.desc()).limit(50).all()
        
        products_list = []
        for fp in foreign_products:
            products_list.append({
                'id': fp.id,
                'name': f"{fp.name} (خارجی - {fp.batch_no})",
                'selling_price': fp.selling_price,
                'buying_price': fp.buying_price,
                'profit_per_item': fp.profit_per_item,
                'unit': fp.unit
            })
        
        return jsonify({
            'success': True,
            'products': products_list
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@main_bp.route('/api/products/barcode/<barcode>')
@login_required
def get_product_by_barcode(barcode):
    """دریافت محصول با بارکد"""
    product = Product.query.filter_by(barcode=barcode).first()
    
    if product:
        return jsonify({
            "id": product.id,
            "name": product.name,
            "selling_price": float(product.selling_price or 0),
            "items_per_carton": product.items_per_carton or 1,
            "total_items": product.total_items,  # ✅ مهم: تعداد کل دانه‌ها
            "quantity": product.quantity,
            "single_quantity": product.single_quantity,
            "unit": product.unit if product.unit else "عدد",
            "barcode": product.barcode
        })
    
    return jsonify(None), 404
    
# ==================== مصارف روزانه ====================
@main_bp.route("/expenses")
@login_required
def list_expenses():
    """لیست مصارف با پشتیبانی از فیلتر تاریخ شمسی"""
    import jdatetime
    from datetime import datetime
    from collections import defaultdict
    
    # دریافت پارامترهای فیلتر
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    
    query = DailyExpense.query
    
    # فیلتر بر اساس تاریخ شمسی
    if start_date_str:
        try:
            parts = start_date_str.split('/')
            if len(parts) == 3:
                jalali_year = int(parts[0])
                jalali_month = int(parts[1])
                jalali_day = int(parts[2])
                jalali_date = jdatetime.date(jalali_year, jalali_month, jalali_day)
                gregorian_date = jalali_date.togregorian()
                start_datetime = datetime.combine(gregorian_date, datetime.min.time())
                query = query.filter(DailyExpense.date >= start_datetime)
                print(f"✅ فیلتر از تاریخ: {start_date_str} -> {start_datetime}")
        except Exception as e:
            print(f"❌ خطا در تبدیل تاریخ شروع: {e}")
    
    if end_date_str:
        try:
            parts = end_date_str.split('/')
            if len(parts) == 3:
                jalali_year = int(parts[0])
                jalali_month = int(parts[1])
                jalali_day = int(parts[2])
                jalali_date = jdatetime.date(jalali_year, jalali_month, jalali_day)
                gregorian_date = jalali_date.togregorian()
                end_datetime = datetime.combine(gregorian_date, datetime.max.time())
                query = query.filter(DailyExpense.date <= end_datetime)
                print(f"✅ فیلتر تا تاریخ: {end_date_str} -> {end_datetime}")
        except Exception as e:
            print(f"❌ خطا در تبدیل تاریخ پایان: {e}")
    
    # دریافت مصارف
    expenses = query.order_by(DailyExpense.date.desc()).all()
    
    # محاسبه مجموع و میانگین
    total_amount = sum(e.amount for e in expenses)
    avg_amount = total_amount / len(expenses) if expenses else 0
    
    # آمار بر اساس دسته‌بندی
    category_stats = defaultdict(float)
    for expense in expenses:
        category_stats[expense.category or 'سایر'] += expense.amount
    
    # ✅ اضافه کردن تاریخ شمسی به هر آیتم (برای نمایش در تمپلیت)
    for expense in expenses:
        if expense.date:
            try:
                jalali = jdatetime.date.fromgregorian(date=expense.date)
                expense.persian_date = jalali.strftime('%Y/%m/%d')
            except Exception as e:
                print(f"❌ خطا در تبدیل تاریخ برای expense {expense.id}: {e}")
                expense.persian_date = expense.date.strftime('%Y-%m-%d') if expense.date else '---'
        else:
            # اگر تاریخ نداشت، از تاریخ ایجاد استفاده کن
            if expense.created_at:
                try:
                    jalali = jdatetime.date.fromgregorian(date=expense.created_at)
                    expense.persian_date = jalali.strftime('%Y/%m/%d')
                except:
                    expense.persian_date = expense.created_at.strftime('%Y-%m-%d') if expense.created_at else '---'
            else:
                expense.persian_date = '---'
    
    print(f"📊 تعداد کل مصارف: {len(expenses)}")
    print(f"💰 مجموع مبلغ: {total_amount:,.0f} افغانی")
    
    return render_template(
        "expenses.html",
        expenses=expenses,
        total_amount=total_amount,
        avg_amount=avg_amount,
        start_date=start_date_str,
        end_date=end_date_str,
        category_stats=dict(category_stats)
    )
    
@main_bp.route("/expenses/add", methods=["GET", "POST"])
@login_required
def add_expense():
    """ثبت مصرف جدید با پشتیبانی از تاریخ شمسی"""
    from datetime import datetime
    import jdatetime
    
    if request.method == "POST":
        try:
            description = request.form.get('description', '').strip()
            amount = float(request.form.get('amount', 0))
            category = request.form.get('category', '').strip()
            jalali_date_str = request.form.get('date', '').strip()
            
            # اعتبارسنجی
            if not description:
                flash("❌ لطفاً توضیحات را وارد کنید", "error")
                return redirect(url_for('main_bp.add_expense'))
                
            if amount <= 0:
                flash("❌ مبلغ باید بزرگتر از صفر باشد", "error")
                return redirect(url_for('main_bp.add_expense'))
            
            if not category:
                flash("❌ لطفاً دسته‌بندی را انتخاب کنید", "error")
                return redirect(url_for('main_bp.add_expense'))
            
            # تبدیل تاریخ شمسی به میلادی
            if jalali_date_str:
                try:
                    # تبدیل تاریخ شمسی به میلادی
                    parts = jalali_date_str.split('/')
                    if len(parts) == 3:
                        jalali_year = int(parts[0])
                        jalali_month = int(parts[1])
                        jalali_day = int(parts[2])
                        
                        jalali_date = jdatetime.date(jalali_year, jalali_month, jalali_day)
                        gregorian_date = jalali_date.togregorian()
                        expense_date = datetime.combine(gregorian_date, datetime.min.time())
                    else:
                        expense_date = datetime.now()
                except Exception as e:
                    print(f"خطا در تبدیل تاریخ: {e}")
                    expense_date = datetime.now()
            else:
                expense_date = datetime.now()
            
            # ایجاد مصرف جدید
            expense = DailyExpense(
                description=description,
                amount=amount,
                category=category,
                date=expense_date,
                created_by=current_user.id
            )
            
            db.session.add(expense)
            db.session.flush()
            
            # کاهش موجودی نقدی
            success = update_cash_balance(
                amount=-amount,
                transaction_type='expense',
                reference_id=expense.id,
                description=f"مصرف: {description} ({category})"
            )
            
            if success:
                db.session.commit()
                flash(f"✅ مصرف '{description}' در دسته {category} به مبلغ {amount:,.0f} افغانی ثبت شد", "success")
                flash(f"💰 مبلغ {amount:,.0f} افغانی از صندوق کسر شد", "info")
            else:
                db.session.rollback()
                flash("❌ خطا در به‌روزرسانی موجودی نقدی", "error")
            
            return redirect(url_for('main_bp.list_expenses'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ خطا در ثبت مصرف: {str(e)}", "error")
            return redirect(url_for('main_bp.add_expense'))
    
    # تاریخ شمسی امروز برای نمایش در فرم
    today_jalali = jdatetime.date.today().strftime('%Y/%m/%d')
    
    return render_template("add_expense.html", now=datetime.now(), today_jalali=today_jalali)

@main_bp.route("/expenses/edit/<int:expense_id>", methods=["GET", "POST"])
@login_required
def edit_expense(expense_id):
    """ویرایش مصرف با پشتیبانی از تاریخ شمسی"""
    import jdatetime
    from datetime import datetime
    
    expense = DailyExpense.query.get_or_404(expense_id)
    
    # اضافه کردن تاریخ شمسی برای نمایش در فرم
    if expense.date:
        try:
            jalali = jdatetime.date.fromgregorian(date=expense.date)
            expense.persian_date = jalali.strftime('%Y/%m/%d')
        except:
            expense.persian_date = ''
    else:
        expense.persian_date = ''
    
    if request.method == "POST":
        try:
            description = request.form.get('description', '').strip()
            amount = float(request.form.get('amount', 0))
            category = request.form.get('category', '').strip()
            persian_date_str = request.form.get('persian_date', '').strip()
            
            # اعتبارسنجی
            if not description:
                flash("❌ لطفاً توضیحات را وارد کنید", "error")
                return redirect(url_for('main_bp.edit_expense', expense_id=expense_id))
            
            if amount <= 0:
                flash("❌ مبلغ باید بزرگتر از صفر باشد", "error")
                return redirect(url_for('main_bp.edit_expense', expense_id=expense_id))
            
            if not category:
                flash("❌ لطفاً دسته‌بندی را انتخاب کنید", "error")
                return redirect(url_for('main_bp.edit_expense', expense_id=expense_id))
            
            # به‌روزرسانی اطلاعات
            expense.description = description
            expense.amount = amount
            expense.category = category
            
            # تبدیل تاریخ شمسی به میلادی
            if persian_date_str:
                try:
                    parts = persian_date_str.split('/')
                    if len(parts) == 3:
                        jalali_year = int(parts[0])
                        jalali_month = int(parts[1])
                        jalali_day = int(parts[2])
                        jalali_date = jdatetime.date(jalali_year, jalali_month, jalali_day)
                        gregorian_date = jalali_date.togregorian()
                        expense.date = datetime.combine(gregorian_date, datetime.min.time())
                except Exception as e:
                    print(f"خطا در تبدیل تاریخ: {e}")
                    # اگر خطا داشت، تاریخ را تغییر نده
            
            db.session.commit()
            flash("✅ مصرف با موفقیت ویرایش شد", "success")
            return redirect(url_for('main_bp.list_expenses'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ خطا در ویرایش مصرف: {str(e)}", "error")
            return redirect(url_for('main_bp.edit_expense', expense_id=expense_id))
    
    return render_template("edit_expense.html", expense=expense)

@main_bp.route("/expenses/delete/<int:expense_id>", methods=["POST"])
@login_required
def delete_expense(expense_id):
    """حذف مصرف"""
    try:
        expense = DailyExpense.query.get_or_404(expense_id)
        description = expense.description
        
        db.session.delete(expense)
        db.session.commit()
        
        flash(f"✅ مصرف '{description}' با موفقیت حذف شد", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"❌ خطا در حذف مصرف: {str(e)}", "error")
    
    return redirect(url_for('main_bp.list_expenses'))

@main_bp.route("/expenses/report")
@login_required
def expenses_report():
    filter_type = request.args.get("filter", "all")
    now = datetime.utcnow()
    
    if filter_type == "today":
        start = datetime(now.year, now.month, now.day)
        expenses = DailyExpense.query.filter(DailyExpense.date >= start).all()
    elif filter_type == "month":
        start = datetime(now.year, now.month, 1)
        expenses = DailyExpense.query.filter(DailyExpense.date >= start).all()
    else:
        expenses = DailyExpense.query.all()
    
    total_expenses = sum(e.amount for e in expenses)
    
    return render_template(
        "expenses_report.html",
        expenses=expenses,
        total_expenses=total_expenses,
        selected_filter=filter_type
    )

@main_bp.route("/expenses/report/print")
@login_required
def expenses_report_print():
    filter_type = request.args.get("filter", "all")
    now = datetime.utcnow()
    
    if filter_type == "today":
        start = datetime(now.year, now.month, now.day)
        expenses = DailyExpense.query.filter(DailyExpense.date >= start).all()
    elif filter_type == "month":
        start = datetime(now.year, now.month, 1)
        expenses = DailyExpense.query.filter(DailyExpense.date >= start).all()
    else:
        expenses = DailyExpense.query.all()
    
    total_expenses = sum(e.amount for e in expenses)
    
    return render_template(
        "expenses_report_print.html",
        expenses=expenses,
        total_expenses=total_expenses,
        selected_filter=filter_type,
        now=now
    )

# ==================== سیستم قرض‌داری ====================
@main_bp.route("/loans")
@login_required
def list_loans():
    loans = Loan.query.order_by(Loan.loan_date.desc()).all()
    total_loans = sum(loan.amount for loan in loans)
    total_paid = sum(loan.amount for loan in loans if loan.is_paid)
    total_remaining = total_loans - total_paid
    
    return render_template("loans.html", 
                         loans=loans, 
                         total_loans=total_loans,
                         total_paid=total_paid,
                         total_remaining=total_remaining)

@main_bp.route("/loans/add", methods=["GET", "POST"])
@login_required
def add_loan():
    form = LoanForm()
    if form.validate_on_submit():
        loan = Loan(
            lender_name=form.lender_name.data,
            amount=form.amount.data,
            description=form.description.data,
            loan_date=form.loan_date.data,
            due_date=form.due_date.data,
            created_by=current_user.id
        )
        db.session.add(loan)
        db.session.commit()
        flash("قرض جدید ثبت شد ✅", "success")
        return redirect(url_for("main.list_loans"))
    
    form.loan_date.data = date.today()
    return render_template("add_loan.html", form=form)

@main_bp.route("/loans/<int:loan_id>/pay", methods=["GET", "POST"])
@login_required
def add_loan_payment(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    form = LoanPaymentForm()
    
    total_paid = sum(payment.amount for payment in loan.payments)
    remaining_amount = loan.amount - total_paid
    
    if form.validate_on_submit():
        amount = form.amount.data
        
        if amount > remaining_amount:
            flash(f"مبلغ پرداختی نمی‌تواند بیشتر از {remaining_amount:,.0f} افغانی باشد ❌", "danger")
            return render_template("add_loan_payment.html", form=form, loan=loan, remaining_amount=remaining_amount)
        
        payment = LoanPayment(
            loan_id=loan.id,
            amount=amount,
            payment_date=form.payment_date.data,
            receipt_number=form.receipt_number.data,
            description=form.description.data,
            created_by=current_user.id
        )
        db.session.add(payment)
        
        new_total_paid = total_paid + amount
        if new_total_paid >= loan.amount:
            loan.is_paid = True
        
        db.session.commit()
        flash("پرداخت قرض ثبت شد ✅", "success")
        return redirect(url_for("main.list_loans"))
    
    form.payment_date.data = date.today()
    form.amount.data = remaining_amount
    
    return render_template("add_loan_payment.html", 
                         form=form, 
                         loan=loan, 
                         remaining_amount=remaining_amount,
                         total_paid=total_paid)

@main_bp.route("/loans/<int:loan_id>/payments")
@login_required
def loan_payments(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    payments = LoanPayment.query.filter_by(loan_id=loan_id).order_by(LoanPayment.payment_date.desc()).all()
    return render_template("loan_payments.html", loan=loan, payments=payments)

# ==================== سیستم استاک اوت ====================
@main_bp.route("/stock/expired")
@login_required
def expired_products():
    today = date.today()
    expired_products = Product.query.filter(
        Product.expiry_date.isnot(None),
        Product.expiry_date < today
    ).order_by(Product.expiry_date.asc()).all()
    
    total_value = sum(p.quantity * p.buying_price for p in expired_products)
    
    return render_template("expired_products.html", 
                         products=expired_products, 
                         today=today,
                         total_value=total_value)

@main_bp.route("/stock/expiring_soon")
@login_required
def expiring_soon_products():
    today = date.today()
    soon_date = today + timedelta(days=30)
    
    expiring_products = Product.query.filter(
        Product.expiry_date.isnot(None),
        Product.expiry_date >= today,
        Product.expiry_date <= soon_date
    ).order_by(Product.expiry_date.asc()).all()
    
    total_value = sum(p.quantity * p.buying_price for p in expiring_products)
    
    return render_template("expiring_soon_products.html", 
                         products=expiring_products, 
                         today=today,
                         soon_date=soon_date,
                         total_value=total_value)
# ==================== سیستم مرجوعی ====================


@main_bp.route('/returns', methods=['GET', 'POST'])
@login_required
def returns():
    form = ReturnProductForm()
    
    # ========== دریافت لیست‌ها برای نمایش در فرم ==========
    sales = Sale.query.order_by(Sale.date.desc()).limit(100).all()
    customers = Customer.query.order_by(Customer.name).all()
    products = Product.query.order_by(Product.name).all()
    
    try:
        # تنظیم choices برای فرم
        form.product_id.choices = [(0, "انتخاب محصول")] + [
            (p.id, f"{p.name} (موجودی: {p.total_items:.0f} {p.unit})") 
            for p in products
        ]
        form.sale_id.choices = [(0, "انتخاب فاکتور")] + [
            (s.id, f"فاکتور {s.invoice_number} - {s.customer.name if s.customer else 'بدون مشتری'} - {s.final_amount:,.0f} افغانی") 
            for s in sales
        ]
        form.customer_id.choices = [(0, "انتخاب مشتری")] + [
            (c.id, f"{c.name} (قرض داری: {c.total_debt:,.0f} افغانی)") 
            for c in customers
        ]
    except Exception as e:
        flash(f"خطا در بارگذاری داده‌ها: {str(e)}", "error")
        return redirect(url_for('main_bp.index'))
    
    if form.validate_on_submit():
        try:
            if form.product_id.data == 0 or form.sale_id.data == 0 or form.customer_id.data == 0:
                flash("لطفاً همه فیلدهای ضروری را انتخاب کنید", "error")
                return render_template('returns.html', form=form, sales=sales, customers=customers, products=products)
            
            # ===== ۱. دریافت اطلاعات =====
            product = Product.query.get(form.product_id.data)
            sale = Sale.query.get(form.sale_id.data)
            customer = Customer.query.get(form.customer_id.data)
            
            if not product or not sale or not customer:
                flash("اطلاعات یافت نشد", "error")
                return render_template('returns.html', form=form, sales=sales, customers=customers, products=products)
            
            # ===== ۲. پیدا کردن آیتم فروش مربوطه =====
            sale_item = SaleItem.query.filter_by(
                sale_id=sale.id, 
                product_id=product.id
            ).first()
            
            if not sale_item:
                flash(f"محصول {product.name} در فاکتور {sale.invoice_number} یافت نشد", "error")
                return render_template('returns.html', form=form, sales=sales, customers=customers, products=products)
            
            # بررسی موجودی کافی برای مرجوعی
            if form.quantity.data > sale_item.quantity:
                flash(f"تعداد مرجوعی ({form.quantity.data}) بیشتر از مقدار فروخته شده ({sale_item.quantity}) است", "error")
                return render_template('returns.html', form=form, sales=sales, customers=customers, products=products)
            
            # ===== ۳. محاسبه مبلغ مرجوعی =====
            # محاسبه مبلغ واقعی این آیتم در فاکتور
            original_total = sale_item.quantity * sale_item.selling_price
            original_discount = original_total * (sale_item.discount_percent / 100)
            original_final = original_total - original_discount
            price_per_unit = original_final / sale_item.quantity
            
            # مبلغ مرجوعی
            refund_amount = form.quantity.data * price_per_unit
            
            # ===== ۴. به‌روزرسانی موجودی انبار =====
            # افزایش موجودی (بر اساس تعداد دانه)
            current_total_items = product.total_items
            product.quantity, product.single_quantity = convert_to_carton_single(
                current_total_items + form.quantity.data,
                product.items_per_carton
            )
            
            print(f"✅ {form.quantity.data} عدد به انبار اضافه شد (موجودی جدید: {product.total_items:.0f})")
            
            # ===== ۵. کاهش قرض داری مشتری =====
            old_customer_debt = customer.total_debt
            customer.total_debt = max(0, customer.total_debt - refund_amount)
            print(f"✅ قرض داری مشتری از {old_customer_debt:,.0f} به {customer.total_debt:,.0f} افغانی کاهش یافت")
            
            # ===== ۶. به‌روزرسانی فاکتور فروش =====
            
            # ۶.۱ کاهش تعداد در آیتم فروش
            if form.quantity.data >= sale_item.quantity:
                # اگر همه مرجوع شد، آیتم را حذف کن
                db.session.delete(sale_item)
            else:
                # اگر بخشی مرجوع شد، تعداد را کم کن
                sale_item.quantity -= form.quantity.data
                
                # محاسبه مجدد مبلغ آیتم
                new_total = sale_item.quantity * sale_item.selling_price
                new_discount = new_total * (sale_item.discount_percent / 100)
                sale_item.final_amount = new_total - new_discount
            
            # ۶.۲ محاسبه مجدد کل فاکتور
            remaining_items = SaleItem.query.filter_by(sale_id=sale.id).all()
            
            if len(remaining_items) == 0:
                # اگر همه آیتم‌ها مرجوع شدند، فاکتور را حذف کن
                db.session.delete(sale)
                flash(f"📄 فاکتور {sale.invoice_number} به دلیل مرجوع کامل حذف شد", "info")
            else:
                # محاسبه مجدد مبالغ فاکتور
                new_total_amount = 0
                new_total_discount = 0
                new_final_amount = 0
                
                for item in remaining_items:
                    item_total = item.quantity * item.selling_price
                    item_discount = item_total * (item.discount_percent / 100)
                    item_final = item_total - item_discount
                    
                    new_total_amount += item_total
                    new_total_discount += item_discount
                    new_final_amount += item_final
                
                # محاسبه سهم این مرجوعی از تخفیف کل فاکتور
                old_total = sale.total_amount
                old_final = sale.final_amount
                
                if old_total > 0:
                    discount_ratio = (old_total - old_final) / old_total
                    # تخفیف اعمال شده روی این مرجوعی
                    item_discount_on_return = refund_amount * discount_ratio
                    # کاهش تخفیف کل فاکتور
                    new_total_discount = max(0, sale.total_discount - item_discount_on_return)
                
                # به‌روزرسانی فاکتور
                sale.total_amount = new_total_amount
                sale.total_discount = new_total_discount
                sale.final_amount = new_final_amount
                
                # محاسبه مجدد قرض داری باقی‌مانده
                total_paid = sale.amount_paid
                if total_paid > new_final_amount:
                    # اگر مبلغ پرداختی بیشتر از مبلغ جدید شد، اضافه پرداخت به عنوان قرض داری منفی ثبت نمی‌شود
                    sale.amount_paid = new_final_amount
                    sale.remaining_debt = 0
                    flash(f"⚠️ مبلغ اضافه پرداخت ({total_paid - new_final_amount:,.0f} افغانی) به عنوان اعتبار مشتری ثبت شد", "info")
                else:
                    sale.remaining_debt = new_final_amount - total_paid
                
                print(f"✅ فاکتور {sale.invoice_number} به‌روزرسانی شد")
                print(f"   مبلغ جدید: {sale.final_amount:,.0f} افغانی")
                print(f"   قرض داری باقی‌مانده: {sale.remaining_debt:,.0f} افغانی")
            
            # ===== ۷. کاهش موجودی نقدی (پرداخت به مشتری) =====
            cash_success = update_cash_balance(
                amount=-refund_amount,
                transaction_type='return',
                reference_id=sale.id if sale.id else None,
                description=f"مرجوعی {product.name} - تعداد: {form.quantity.data} عدد - دلیل: {form.reason.data}"
            )
            
            if not cash_success:
                flash("⚠️ مرجوعی ثبت شد اما خطا در به‌روزرسانی موجودی نقدی", "warning")
            else:
                print(f"💰 مبلغ {refund_amount:,.0f} افغانی از صندوق کسر شد")
            
            # ===== ۸. ثبت مرجوعی =====
            return_product = ReturnProduct(
                product_id=form.product_id.data,
                sale_id=form.sale_id.data,
                customer_id=form.customer_id.data,
                quantity=form.quantity.data,
                reason=form.reason.data,
                refund_amount=refund_amount,
                return_date=datetime.utcnow(),
                created_by=current_user.id
            )
            
            db.session.add(return_product)
            db.session.commit()
            
            flash('✅ مرجوعی با موفقیت ثبت شد', 'success')
            flash(f'💰 مبلغ {refund_amount:,.0f} افغانی به مشتری بازگردانده شد', 'info')
            flash(f'📦 {form.quantity.data} عدد {product.unit} به انبار اضافه شد', 'info')
            
            return redirect(url_for('main_bp.returns_list'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ خطا در ثبت مرجوعی: {str(e)}', 'error')
            import traceback
            traceback.print_exc()
    
    # ========== ارسال متغیرها به تمپلیت ==========
    return render_template(
        'returns.html',
        form=form,
        sales=sales,
        customers=customers,
        products=products
    )

@main_bp.route('/returns/list')
@login_required
def returns_list():
    try:
        returns = ReturnProduct.query.order_by(ReturnProduct.return_date.desc()).all()
        
        # محاسبه آمار
        total_refund = sum(r.refund_amount for r in returns)
        total_items = sum(r.quantity for r in returns)
        
        return render_template(
            'returns_list.html', 
            returns=returns,
            total_refund=total_refund,
            total_items=total_items
        )
    except Exception as e:
        flash(f'❌ خطا در بارگذاری لیست مرجوعی‌ها: {str(e)}', 'error')
        return redirect(url_for('main_bp.index'))


@main_bp.route('/returns/delete/<int:id>', methods=['POST'])
@login_required
def delete_return(id):
    try:
        return_product = ReturnProduct.query.get_or_404(id)
        sale = Sale.query.get(return_product.sale_id)
        product = Product.query.get(return_product.product_id)
        customer = Customer.query.get(return_product.customer_id)
        
        # ===== بازگردانی تغییرات (Undo) =====
        if product:
            # کم کردن از انبار
            current_total_items = product.total_items
            product.quantity, product.single_quantity = convert_to_carton_single(
                max(0, current_total_items - return_product.quantity),
                product.items_per_carton
            )
        
        if customer:
            # برگرداندن قرض داری
            customer.total_debt += return_product.refund_amount
        
        if sale:
            # بازگردانی آیتم فروش
            sale_item = SaleItem.query.filter_by(
                sale_id=sale.id, 
                product_id=return_product.product_id
            ).first()
            
            if sale_item:
                # افزایش تعداد در آیتم موجود
                sale_item.quantity += return_product.quantity
                sale_item.final_amount = (sale_item.quantity * sale_item.selling_price) * (1 - sale_item.discount_percent / 100)
            else:
                # بازگردانی آیتم حذف شده
                # باید اطلاعات اصلی از مرجوعی ذخیره شده باشد
                # در اینجا یک رکورد موقت ایجاد می‌کنیم
                sale_item = SaleItem(
                    sale_id=sale.id,
                    product_id=return_product.product_id,
                    quantity=return_product.quantity,
                    selling_price=product.selling_price if product else 0,
                    discount_percent=0,
                    final_amount=return_product.quantity * (product.selling_price if product else 0)
                )
                db.session.add(sale_item)
            
            # محاسبه مجدد مبالغ فاکتور
            total_amount = 0
            total_discount = 0
            final_amount = 0
            
            items = SaleItem.query.filter_by(sale_id=sale.id).all()
            for item in items:
                item_total = item.quantity * item.selling_price
                item_discount = item_total * (item.discount_percent / 100)
                item_final = item_total - item_discount
                
                total_amount += item_total
                total_discount += item_discount
                final_amount += item_final
            
            sale.total_amount = total_amount
            sale.total_discount = total_discount
            sale.final_amount = final_amount
            
            # محاسبه مجدد قرض داری
            if sale.amount_paid > final_amount:
                sale.amount_paid = final_amount
            sale.remaining_debt = final_amount - sale.amount_paid
        
        # برگرداندن پول به صندوق
        update_cash_balance(
            amount=return_product.refund_amount,
            transaction_type='adjustment',
            description=f"حذف مرجوعی {return_product.product.name if return_product.product else 'محصول'}"
        )
        
        # حذف مرجوعی
        db.session.delete(return_product)
        db.session.commit()
        
        flash('🗑️ مرجوعی حذف شد و تغییرات بازگردانده شد', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ خطا در حذف مرجوعی: {str(e)}', 'error')
    
    return redirect(url_for('main_bp.returns_list'))


# ==================== تابع کمکی ====================
def convert_to_carton_single(total_items, items_per_carton):
    """تبدیل تعداد کل دانه به کارتن و تک"""
    if items_per_carton <= 0:
        items_per_carton = 1
    cartons = total_items // items_per_carton
    singles = total_items % items_per_carton
    return cartons, singles

# ===== API برای دریافت اطلاعات فاکتور (برای فرم مرجوعی) =====
@main_bp.route('/api/sale/<int:sale_id>/info')
@login_required
def get_sale_info(sale_id):
    """دریافت اطلاعات فاکتور برای مرجوعی"""
    sale = Sale.query.get_or_404(sale_id)
    
    items = []
    for item in sale.items:
        items.append({
            'product_id': item.product_id,
            'product_name': item.product.name if item.product else 'محصول خارجی',
            'quantity': item.quantity,
            'selling_price': item.selling_price,
            'discount_percent': item.discount_percent,
            'final_amount': item.final_amount,
            'unit': item.product.unit if item.product else 'عدد'
        })
    
    return jsonify({
        'success': True,
        'sale': {
            'id': sale.id,
            'invoice_number': sale.invoice_number,
            'customer_id': sale.customer_id,
            'customer_name': sale.customer.name if sale.customer else 'مشتری عمومی',
            'total_amount': sale.total_amount,
            'final_amount': sale.final_amount,
            'remaining_debt': sale.remaining_debt,
            'items': items
        }
    })
  
@main_bp.route("/profit_report")
@login_required
def profit_report():
    """گزارش سود و زیان با محاسبه دقیق به افغانی و پشتیبانی از سیستم کارتن و تک"""
    from datetime import datetime, timedelta
    import time
    import jdatetime
    from app.models import CashWithdrawal  # ✅ اضافه کنید
    
    start_time = time.time()
    
    # ========== دریافت پارامترهای فیلتر ==========
    filter_type = request.args.get('filter', 'month')
    selected_month = request.args.get('month', type=int)
    
    # ========== تاریخ شمسی امروز ==========
    today_jalali = jdatetime.date.today()
    today_gregorian = datetime.now().date()
    
    # ========== لیست ماه‌های دری (افغانستان) ==========
    month_names = ['حمل', 'ثور', 'جوزا', 'سرطان', 'اسد', 'سنبله', 
                   'میزان', 'عقرب', 'قوس', 'جدی', 'دلو', 'حوت']
    
    def get_gregorian_range_from_jalali_month(jalali_year, jalali_month):
        first_day_jalali = jdatetime.date(jalali_year, jalali_month, 1)
        if jalali_month < 12:
            last_day_jalali = jdatetime.date(jalali_year, jalali_month + 1, 1) - timedelta(days=1)
        else:
            last_day_jalali = jdatetime.date(jalali_year + 1, 1, 1) - timedelta(days=1)
        
        start_date = first_day_jalali.togregorian()
        end_date = last_day_jalali.togregorian()
        
        return start_date, end_date
    
    # ========== اعمال فیلتر ==========
    start_date = None
    end_date = today_gregorian
    
    if filter_type == 'today':
        start_date = today_gregorian
        end_date = today_gregorian
    
    elif filter_type == 'week':
        week_day = today_jalali.weekday()
        start_of_week_jalali = today_jalali - timedelta(days=week_day)
        end_of_week_jalali = start_of_week_jalali + timedelta(days=6)
        start_date = start_of_week_jalali.togregorian()
        end_date = end_of_week_jalali.togregorian()
    
    elif filter_type == 'month':
        if selected_month and 1 <= selected_month <= 12:
            jalali_year = today_jalali.year
            if selected_month > today_jalali.month:
                jalali_year -= 1
            start_date, end_date = get_gregorian_range_from_jalali_month(jalali_year, selected_month)
        else:
            start_date, end_date = get_gregorian_range_from_jalali_month(today_jalali.year, today_jalali.month)
    
    elif filter_type == 'all':
        start_date = datetime(2020, 1, 1).date()
        end_date = today_gregorian
    
    else:
        start_date, end_date = get_gregorian_range_from_jalali_month(today_jalali.year, today_jalali.month)
    
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    # ========== محاسبات فروش ==========
    sales_query = Sale.query.filter(Sale.date >= start_datetime, Sale.date <= end_datetime)
    sales = sales_query.all()
    
    total_sales = sum(s.final_amount or 0 for s in sales)
    sales_count = len(sales)
    total_discounts = sum(s.total_discount or 0 for s in sales)
    
    # ========== محاسبه بهای تمام شده کالای فروش رفته و سود ناخالص ==========
    total_cogs = 0
    total_profit_from_sales = 0
    
    for sale in sales:
        for item in sale.items:
            if item.product:
                product = item.product
                quantity = item.quantity
                cost = (product.buying_price or 0) * quantity
                total_cogs += cost
                profit = ((item.selling_price or 0) - (product.buying_price or 0)) * quantity
                total_profit_from_sales += profit
                
            elif item.foreign_name and item.foreign_selling_price:
                estimated_buying_price = item.foreign_selling_price * 0.7
                cost = estimated_buying_price * item.quantity
                total_cogs += cost
                profit = (item.foreign_selling_price - estimated_buying_price) * item.quantity
                total_profit_from_sales += profit
    
    # ========== محاسبه مرجوعی‌ها ==========
    total_returns_refund = 0
    total_cogs_returned = 0
    total_profit_loss_from_returns = 0
    
    returns_query = ReturnProduct.query.filter(
        ReturnProduct.return_date >= start_datetime,
        ReturnProduct.return_date <= end_datetime
    ).all()
    
    for return_item in returns_query:
        total_returns_refund += return_item.refund_amount or 0
        
        if return_item.product:
            product = return_item.product
            cogs_returned = (product.buying_price or 0) * return_item.quantity
            total_cogs_returned += cogs_returned
            profit_loss = (return_item.refund_amount or 0) - cogs_returned
            total_profit_loss_from_returns += max(0, profit_loss)
        else:
            estimated_buying = (return_item.refund_amount or 0) * 0.7
            total_cogs_returned += estimated_buying
            profit_loss = (return_item.refund_amount or 0) - estimated_buying
            total_profit_loss_from_returns += max(0, profit_loss)
    
    returns_count = len(returns_query)
    
    # ========== ✅ بخش جدید: محاسبه برداشت‌های شخصی ==========
    withdrawals_query = CashWithdrawal.query.filter(
        CashWithdrawal.withdrawal_date >= start_datetime,
        CashWithdrawal.withdrawal_date <= end_datetime
    ).all()
    
    total_withdrawals = sum(w.amount or 0 for w in withdrawals_query)
    withdrawals_count = len(withdrawals_query)
    
    # گروه‌بندی برداشت‌ها بر اساس هدف و کاربر
    withdrawals_by_purpose = {}
    withdrawals_by_user = {}
    
    for w in withdrawals_query:
        # گروه‌بندی بر اساس هدف
        purpose = w.purpose or 'سایر'
        withdrawals_by_purpose[purpose] = withdrawals_by_purpose.get(purpose, 0) + w.amount
        
        # گروه‌بندی بر اساس کاربر
        user_name = w.employee.full_name if w.employee else 'نامشخص'
        if user_name not in withdrawals_by_user:
            withdrawals_by_user[user_name] = {'amount': 0, 'count': 0, 'details': []}
        withdrawals_by_user[user_name]['amount'] += w.amount
        withdrawals_by_user[user_name]['count'] += 1
        withdrawals_by_user[user_name]['details'].append({
            'amount': w.amount,
            'purpose': w.purpose,
            'date': w.withdrawal_date.strftime('%Y-%m-%d %H:%M') if w.withdrawal_date else 'نامشخص',
            'description': w.description or ''
        })
    
    # ========== محاسبه ارزش موجودی کالا ==========
    products = Product.query.all()
    total_inventory_value_end = 0
    total_inventory_cost_end = 0
    total_items_count_end = 0
    
    for product in products:
        total_items = product.total_items
        if total_items > 0:
            total_items_count_end += total_items
            total_inventory_value_end += total_items * (product.selling_price or 0)
            total_inventory_cost_end += total_items * (product.buying_price or 0)
    
    # ========== محاسبات هزینه‌های عملیاتی ==========
    expenses_query = DailyExpense.query.filter(
        DailyExpense.date >= start_datetime,
        DailyExpense.date <= end_datetime
    )
    total_operating_expenses = sum(e.amount or 0 for e in expenses_query.all())
    expenses_count = expenses_query.count()
    
    # ضرر و زیان موجودی
    losses_query = InventoryLoss.query.filter(
        InventoryLoss.loss_date >= start_datetime,
        InventoryLoss.loss_date <= end_datetime
    )
    total_inventory_loss = sum(l.total_loss or 0 for l in losses_query.all())
    losses_count = losses_query.count()
    
    # پرداخت اقساط قرض‌ها
    loan_payments_query = LoanPayment.query.filter(
        LoanPayment.payment_date >= start_date,
        LoanPayment.payment_date <= end_date
    )
    total_loan_interest = sum(lp.amount or 0 for lp in loan_payments_query.all())
    loan_payments_count = loan_payments_query.count()
    
    # ========== محاسبات نهایی سود و زیان ==========
    gross_profit = total_sales - total_cogs
    operating_profit = gross_profit - total_operating_expenses - total_inventory_loss
    net_profit = operating_profit - total_loan_interest
    
    # سود خالص بعد از برداشت‌ها (برای نمایش در گزارش)
    net_profit_after_withdrawals = net_profit - total_withdrawals
    
    gross_profit_margin = (gross_profit / total_sales * 100) if total_sales > 0 else 0
    net_profit_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
    
    # ========== سایر اطلاعات مالی ==========
    loans_received_query = Loan.query.filter(
        Loan.loan_date >= start_date,
        Loan.loan_date <= end_date
    )
    total_loans_received = sum(l.amount or 0 for l in loans_received_query.all())
    loans_received_count = loans_received_query.count()
    
    cash = CashBalance.query.first()
    cash_balance = cash.amount if cash else 0
    
    total_customer_debt = db.session.query(db.func.sum(Customer.total_debt)).scalar() or 0
    total_creditor_debt = db.session.query(db.func.sum(Creditor.current_debt)).scalar() or 0
    net_debt_position = total_creditor_debt - total_customer_debt
    
    total_assets = cash_balance + total_inventory_cost_end + total_customer_debt
    
    # ========== داده‌های نمودار ==========
    chart_labels = ['فروش', 'سود ناخالص', 'سود عملیاتی', 'سود خالص']
    chart_data = [float(total_sales), float(gross_profit), float(operating_profit), float(net_profit)]
    
    # نمودار فروش روزانه
    daily_sales_data = []
    daily_sales_labels = []
    
    if (end_date - start_date).days <= 31:
        current_date = start_date
        while current_date <= end_date:
            day_start = datetime.combine(current_date, datetime.min.time())
            day_end = datetime.combine(current_date, datetime.max.time())
            
            day_sales = db.session.query(db.func.sum(Sale.final_amount)).filter(
                Sale.date >= day_start,
                Sale.date <= day_end
            ).scalar() or 0
        
            daily_sales_labels.append(current_date.strftime('%Y-%m-%d'))
            daily_sales_data.append(float(day_sales))
            
            current_date += timedelta(days=1)
    
    # نمودار هزینه‌ها بر اساس دسته‌بندی
    expenses_data = []
    expenses_labels = []
    
    if hasattr(DailyExpense, 'category'):
        expenses_by_category = db.session.query(
            DailyExpense.category,
            db.func.sum(DailyExpense.amount)
        ).filter(
            DailyExpense.date >= start_datetime,
            DailyExpense.date <= end_datetime
        ).group_by(DailyExpense.category).all()
        
        if expenses_by_category:
            expenses_labels = [item[0] or 'سایر' for item in expenses_by_category]
            expenses_data = [float(item[1]) for item in expenses_by_category]
    
    if not expenses_data:
        expenses_labels = ['هزینه‌های عملیاتی']
        expenses_data = [float(total_operating_expenses)] if total_operating_expenses > 0 else [0]
    
    # ========== لیست ماه‌ها ==========
    months_list = [(i+1, name) for i, name in enumerate(month_names)]
    selected_months = [selected_month] if selected_month else []
    selected_months_names = [month_names[selected_month-1]] if selected_month and 1 <= selected_month <= 12 else []
    
    # ========== زمان اجرا ==========
    execution_time = time.time() - start_time
    
    return render_template(
        "profit_report.html",
        # فیلترها
        selected_filter=filter_type,
        selected_months=selected_months,
        selected_months_names=selected_months_names,
        months_list=months_list,
        start_date=start_datetime,
        end_date=end_datetime,
        
        # فروش
        total_sales=total_sales,
        sales_count=sales_count,
        total_discounts=total_discounts,
        
        # سودها
        gross_profit=gross_profit,
        operating_profit=operating_profit,
        net_profit=net_profit,
        net_profit_after_withdrawals=net_profit_after_withdrawals,  # ✅ جدید
        
        # بهای تمام شده
        total_cost_of_goods_sold=total_cogs,
        total_expenses=total_operating_expenses,
        total_inventory_loss=total_inventory_loss,
        
        # مرجوعی‌ها
        total_returns_amount=total_returns_refund,
        total_profit_from_returns=total_profit_loss_from_returns,
        returns_count=returns_count,
        
        # ✅ بخش برداشت‌ها (جدید)
        total_withdrawals=total_withdrawals,
        withdrawals_count=withdrawals_count,
        withdrawals_by_purpose=withdrawals_by_purpose,
        withdrawals_by_user=withdrawals_by_user,
        withdrawals_list=withdrawals_query,
        
        # سایر هزینه‌ها
        total_loan_payments=total_loan_interest,
        total_loans_received=total_loans_received,
        
        # درصدها
        gross_profit_margin=gross_profit_margin,
        net_profit_margin=net_profit_margin,
        
        # موجودی نقدی
        cash_balance=cash_balance,
        
        # ارزش موجودی
        total_inventory_value=total_inventory_value_end,
        total_inventory_cost=total_inventory_cost_end,
        total_items_count=total_items_count_end,
        
        # بدهی‌ها
        total_customer_debt=total_customer_debt,
        total_creditor_debt=total_creditor_debt,
        net_debt_position=net_debt_position,
        
        # کل دارایی‌ها
        total_assets=total_assets,
        
        # نمودارها
        chart_labels=chart_labels,
        chart_data=chart_data,
        daily_sales_data=daily_sales_data,
        daily_sales_labels=daily_sales_labels,
        expenses_data=expenses_data,
        expenses_labels=expenses_labels,
        
        # زمان اجرا
        execution_time=execution_time
    )
@main_bp.route("/report/low_stock")
@login_required
def low_stock_report():
    threshold = request.args.get("threshold", 5, type=int)
    products = Product.query.filter(Product.quantity < threshold).order_by(Product.quantity.asc()).all()
    return render_template("low_stock_report.html", products=products, threshold=threshold)

@main_bp.route("/report/debt_payments")
@login_required
def debt_payments_report():
    payments = DebtPayment.query.order_by(DebtPayment.date.desc()).all()
    return render_template("debt_payments_report.html", payments=payments)

@main_bp.route("/report/user_activities")
@login_required
def user_activities_report():
    try:
        user_sales = db.session.query(
            User.full_name,
            db.func.count(Sale.id).label('sales_count'),
            db.func.sum(Sale.total_amount).label('total_sales')
        ).join(Sale, User.id == Sale.created_by)\
         .group_by(User.id, User.full_name).all()
        
        user_expenses = db.session.query(
            User.full_name,
            db.func.count(DailyExpense.id).label('expenses_count'),
            db.func.sum(DailyExpense.amount).label('total_expenses')
        ).join(DailyExpense, User.id == DailyExpense.created_by)\
         .group_by(User.id, User.full_name).all()
        
        user_payments = db.session.query(
            User.full_name,
            db.func.count(DebtPayment.id).label('payments_count'),
            db.func.sum(DebtPayment.amount).label('total_payments')
        ).join(DebtPayment, User.id == DebtPayment.created_by)\
         .group_by(User.id, User.full_name).all()

        return render_template(
            "user_activities_report.html",
            user_sales=user_sales,
            user_expenses=user_expenses,
            user_payments=user_payments,
            now=datetime.utcnow()
        )
    except Exception as e:
        flash("خطا در بارگذاری گزارش", "danger")
        return redirect(url_for("main_bp.index"))

# ==================== مدیریت کاربران ====================
@main_bp.route("/users")
@login_required
def list_users():
    users = User.query.all()
    return render_template("users.html", users=users)

@main_bp.route("/users/add", methods=["GET", "POST"])
@login_required
def add_user():
    form = UserForm()
    if form.validate_on_submit():
        existing_user = User.query.filter_by(username=form.username.data).first()
        if existing_user:
            flash("این نام کاربری قبلاً ثبت شده است", "danger")
            return render_template("add_user.html", form=form)
        
        user = User(
            username=form.username.data,
            full_name=form.full_name.data,
            role=form.role.data
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash("کاربر جدید با موفقیت ثبت شد ✅", "success")
        return redirect(url_for("main.list_users"))
    
    return render_template("add_user.html", form=form)

# ==================== API ها ====================
@main_bp.route("/api/search_customers_advanced")
@login_required
def api_search_customers_advanced():
    q = request.args.get("q", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 20
    
    if q:
        customers = Customer.query.filter(
            db.or_(
                Customer.name.ilike(f"%{q}%"),
                Customer.phone.ilike(f"%{q}%")
            )
        ).order_by(Customer.name).paginate(page=page, per_page=per_page, error_out=False)
    else:
        customers = Customer.query.order_by(Customer.name).paginate(page=page, per_page=per_page, error_out=False)
    
    results = {
        "results": [{"id": c.id, "text": f"{c.name} ({c.phone or 'بدون شماره'}) - قرض داری: {c.total_debt:,.0f} افغانی"} for c in customers.items],
        "pagination": {
            "more": customers.has_next
        }
    }
    return jsonify(results)

@main_bp.route("/customers/search")
@login_required
def customer_search():
    return render_template("customer_search.html")

@main_bp.route("/api/customers/search")
@login_required
def api_customer_search():
    name = request.args.get("name", "").strip()
    phone = request.args.get("phone", "").strip()
    debt_status = request.args.get("debt_status", "all")
    page = request.args.get("page", 1, type=int)
    per_page = 20

    query = Customer.query

    if name:
        query = query.filter(Customer.name.ilike(f"%{name}%"))

    if phone:
        query = query.filter(Customer.phone.ilike(f"%{phone}%"))

    if debt_status == "debtor":
        query = query.filter(Customer.total_debt > 0)
    elif debt_status == "no_debt":
        query = query.filter(Customer.total_debt == 0)

    customers_pagination = query.order_by(Customer.name).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        "customers": [
            {
                "id": c.id,
                "name": c.name,
                "phone": c.phone,
                "total_debt": c.total_debt,
                "created_at": c.created_at.isoformat()
            }
            for c in customers_pagination.items
        ],
        "total": customers_pagination.total,
        "page": customers_pagination.page,
        "pages": customers_pagination.pages,
        "has_prev": customers_pagination.has_prev,
        "has_next": customers_pagination.has_next,
        "prev_num": customers_pagination.prev_num,
        "next_num": customers_pagination.next_num
    })

@main_bp.route("/api/add_quick_customer", methods=["POST"])
@login_required
def api_add_quick_customer():
    try:
        data = request.get_json()
        
        existing_customer = Customer.query.filter_by(name=data['name']).first()
        if existing_customer:
            return jsonify({
                'success': False,
                'message': 'مشتری با این نام از قبل وجود دارد'
            })
        
        customer = Customer(
            name=data['name'],
            phone=data.get('phone', ''),
            total_debt=0.0
        )
        
        db.session.add(customer)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'customer': {
                'id': customer.id,
                'name': customer.name,
                'phone': customer.phone
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        })

@main_bp.route("/api/add_quick_product", methods=["POST"])
@login_required
def api_add_quick_product():
    try:
        data = request.get_json()
        
        product = Product(
            name=data['name'],
            quantity=0,
            unit=data.get('unit', 'عدد'),
            buying_price=float(data.get('buying_price', 0)),
            selling_price=float(data['selling_price']),
            batch_no=data.get('batch_no', ''),
        )
        
        db.session.add(product)
        db.session.flush()
        
        return jsonify({
            'success': True,
            'product': {
                'id': product.id,
                'name': product.name,
                'selling_price': product.selling_price,
                'quantity': product.quantity,
                'unit': product.unit
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        })

# ==================== بکاپ CSV ====================
@main_bp.route("/backup/export")
@login_required
def backup_export():
    si = io.StringIO()
    w = csv.writer(si)
    w.writerow(["type", "id", "data"])

    try:
        for p in Product.query.all():
            w.writerow(["product", p.id, f"{p.name}|{p.quantity}|{p.buying_price}|{p.selling_price}|{p.unit}|{p.expiry_date}"])

        for c in Customer.query.all():
            w.writerow(["customer", c.id, f"{c.name}|{c.phone}|{c.total_debt}"])

        for s in Sale.query.all():
            w.writerow(["sale", s.id, f"{s.invoice_number}|{s.customer_id}|{s.total_amount}|{s.total_discount}|{s.amount_paid}|{s.remaining_debt}|{s.date}"])
            for it in s.items:
                w.writerow(["sale_item", it.id, f"{s.id}|{it.product_id}|{it.quantity}|{it.selling_price}|{it.discount_percent}"])

        for p in DebtPayment.query.all():
            w.writerow(["debt_payment", p.id, f"{p.customer_id}|{p.amount}|{p.date}|{p.receipt_number}"])

        for e in DailyExpense.query.all():
            w.writerow(["expense", e.id, f"{e.description}|{e.amount}|{e.date}"])

        for r in ReturnProduct.query.all():
            w.writerow(["return_product", r.id, f"{r.product_id}|{r.sale_id}|{r.customer_id}|{r.quantity}|{r.refund_amount}|{r.reason}|{r.return_date}"])

        for l in Loan.query.all():
            w.writerow(["loan", l.id, f"{l.lender_name}|{l.amount}|{l.loan_date}|{l.due_date}|{l.is_paid}"])

        for lp in LoanPayment.query.all():
            w.writerow(["loan_payment", lp.id, f"{lp.loan_id}|{lp.amount}|{lp.payment_date}|{lp.receipt_number}"])

        output = io.BytesIO()
        output.write(si.getvalue().encode("utf-8"))
        output.seek(0)
        
        flash("پشتیبان‌گیری با موفقیت انجام شد ✅", "success")
        return send_file(output, mimetype="text/csv", as_attachment=True, download_name=f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")

    except Exception as e:
        flash(f"خطا در پشتیبان‌گیری: {str(e)} ❌", "error")
        return redirect(url_for('main_bp.index'))

@main_bp.route("/customer/<int:customer_id>/full_report")
@login_required
def customer_full_report(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    
    sales = Sale.query.filter_by(customer_id=customer.id).order_by(Sale.date.desc()).all()
    payments = DebtPayment.query.filter_by(customer_id=customer.id).order_by(DebtPayment.date.desc()).all()
    
    total_sales_amount = sum(sale.total_amount for sale in sales)
    total_paid = sum(sale.amount_paid for sale in sales)
    total_debt_payments = sum(payment.amount for payment in payments)
    current_debt = customer.total_debt or 0
    total_remaining_debt = sum(sale.remaining_debt for sale in sales)
    
    return render_template(
        "customer_full_report.html",
        customer=customer,
        sales=sales,
        payments=payments,
        total_sales_amount=total_sales_amount,
        total_paid=total_paid,
        total_debt_payments=total_debt_payments,
        current_debt=current_debt,
        total_remaining_debt=total_remaining_debt,
        now=datetime.utcnow()
    )
    
@main_bp.route("/backup/manage")
@login_required
def backup_manage():
    """صفحه مدیریت بک‌آپ"""
    # محاسبه آمار
    from app.models import Product, Customer
    
    total_products = Product.query.count()
    total_customers = Customer.query.count()
    
    # محاسبه ارزش موجودی
    total_inventory = 0
    for product in Product.query.all():
        total_inventory += product.selling_price * product.quantity
    
    return render_template(
        "backup_management.html",
        total_products=total_products,
        total_customers=total_customers,
        total_inventory=total_inventory
    )

@main_bp.route('/backup/restore', methods=['GET', 'POST'])
@login_required
@csrf.exempt
def backup_restore():
    if request.method == 'POST':

        if 'backup_file' not in request.files:
            flash('لطفاً فایل پشتیبان را انتخاب کنید', 'error')
            return redirect(url_for('main_bp.backup_manage'))

        file = request.files['backup_file']

        if not file or file.filename == '':
            flash('فایل انتخاب نشده است', 'error')
            return redirect(url_for('main_bp.backup_manage'))

        if not file.filename.lower().endswith('.db'):
            flash('فقط فایل‌های دیتابیس (.db) مجاز هستند', 'error')
            return redirect(url_for('main_bp.backup_manage'))

        try:
            from flask import current_app
            import os, shutil, datetime

            # مسیر دیتابیس واقعی Flask
            db_path = current_app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')

            instance_dir = os.path.dirname(db_path)
            os.makedirs(instance_dir, exist_ok=True)

            # ذخیره موقت فایل آپلودی
            upload_path = os.path.join(instance_dir, secure_filename(file.filename))
            file.save(upload_path)

            # بکاپ قبل از restore
            backup_dir = os.path.join(instance_dir, 'restore_backups')
            os.makedirs(backup_dir, exist_ok=True)

            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_before_restore = os.path.join(
                backup_dir, f'before_restore_{timestamp}.db'
            )

            if os.path.exists(db_path):
                shutil.copy2(db_path, backup_before_restore)

            # بستن session قبل از جایگزینی
            from extensions import db
            db.session.close()

            # جایگزینی دیتابیس
            shutil.copy2(upload_path, db_path)

            flash('✅ دیتابیس با موفقیت بازیابی شد. لطفاً برنامه را ری‌استارت کنید.', 'success')
            return redirect(url_for('main_bp.backup_manage'))

        except Exception as e:
            flash(f'❌ خطا در بازیابی: {e}', 'error')
            return redirect(url_for('main_bp.backup_manage'))

    return render_template('backup_restore.html', title='بازیابی پشتیبان')


# ==================== بکاپ کامل (جدید) ====================
@main_bp.route("/backup/full")
@login_required
def backup_full():
    """بک‌آپ کامل به صورت ZIP"""
    try:
        import zipfile, json, io, os
        from datetime import datetime
        
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # ۱. دیتابیس
            db_file = 'supermarket.db'
            if os.path.exists(db_file):
                zip_file.write(db_file, 'database.db')
            
            # ۲. داده‌های محصولات
            products_data = []
            for product in Product.query.all():
                product_data = {
                    'id': product.id,
                    'name': product.name,
                    'quantity': product.quantity,
                    'unit': product.unit,
                    'buying_price': product.buying_price,
                    'selling_price': product.selling_price,
                    'created_at': product.created_at.isoformat() if product.created_at else None,
                    'updated_at': product.updated_at.isoformat() if product.updated_at else None
                }
                
                # اضافه کردن فیلدهای اختیاری
                if hasattr(product, 'batch_no') and product.batch_no:
                    product_data['batch_no'] = product.batch_no
                else:
                    product_data['batch_no'] = None  # یا رشته خالی
                
                if hasattr(product, 'expiry_date') and product.expiry_date:
                    product_data['expiry_date'] = product.expiry_date.isoformat()
                
                if hasattr(product, 'purchase_type') and product.purchase_type:
                    product_data['purchase_type'] = product.purchase_type
                
                if hasattr(product, 'creditor_id') and product.creditor_id:
                    product_data['creditor_id'] = product.creditor_id
                
                if hasattr(product, 'purchase_description') and product.purchase_description:
                    product_data['purchase_description'] = product.purchase_description
                
                products_data.append(product_data)
            
            # ۳. داده‌های مشتریان
            customers_data = []
            for customer in Customer.query.all():
                customer_data = {
                    'id': customer.id,
                    'name': customer.name,
                    'created_at': customer.created_at.isoformat() if customer.created_at else None,
                    'updated_at': customer.updated_at.isoformat() if customer.updated_at else None
                }
                
                if hasattr(customer, 'phone') and customer.phone:
                    customer_data['phone'] = customer.phone
                
                if hasattr(customer, 'address') and customer.address:
                    customer_data['address'] = customer.address
                
                if hasattr(customer, 'total_debt') and customer.total_debt is not None:
                    customer_data['total_debt'] = customer.total_debt
                
                customers_data.append(customer_data)
            
            # ساختار JSON
            backup_data = {
                'backup_time': datetime.now().isoformat(),
                'products': products_data,
                'customers': customers_data,
                'stats': {
                    'product_count': len(products_data),
                    'customer_count': len(customers_data),
                    'total_inventory_value': sum(p.get('selling_price', 0) * p.get('quantity', 0) 
                                                for p in products_data)
                }
            }
            
            zip_file.writestr('data.json', json.dumps(backup_data, indent=2, ensure_ascii=False))
            
            # ۴. فایل اطلاعات
            info = f"""
            پشتیبان‌گیری فروشگاه
            =====================
            تاریخ: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            تعداد محصولات: {len(products_data)}
            تعداد مشتریان: {len(customers_data)}
            
            فیلدهای محصولات:
            - نام: {all('name' in p for p in products_data)}
            - تعداد: {all('quantity' in p for p in products_data)}
            - شماره بچ: {all('batch_no' in p for p in products_data)}
            - تاریخ انقضا: {sum(1 for p in products_data if p.get('expiry_date'))}
            
            نحوه استفاده:
            ۱. این فایل ZIP را در جای امن نگهداری کنید
            ۲. برای بازگردانی، فایل database.db را جایگزین کنید
            ۳. فایل data.json برای مشاهده داده‌ها
            
            توجه: این فایل حاوی اطلاعات حساس است.
            """
            zip_file.writestr('README.txt', info)
        
        zip_buffer.seek(0)
        filename = f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        
        flash(f"✅ بک‌آپ کامل ایجاد شد: {len(products_data)} محصول", "success")
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        print(f"خطای کامل: {traceback.format_exc()}")
        flash(f"❌ خطا در ایجاد بک‌آپ: {str(e)}", "error")
        return redirect(url_for('main_bp.index'))    
# ==================== سیستم استاک‌اوت ====================
@main_bp.route("/stock/write_off", methods=["GET", "POST"])
@login_required
def inventory_write_off():
    form = InventoryWriteOffForm()
    
    # پر کردن لیست محصولات
    form.product_id.choices = [(0, "انتخاب محصول")] + [
        (p.id, f"{p.name} (موجودی: {p.quantity} {p.unit} - قیمت خرید: {p.buying_price:,.0f} افغانی)") 
        for p in Product.query.filter(Product.quantity > 0).order_by(Product.name).all()
    ]
    
    if form.validate_on_submit():
        try:
            if form.product_id.data == 0:
                flash("❌ لطفاً یک محصول انتخاب کنید", "danger")
                return render_template("inventory_write_off.html", form=form)
            
            product = Product.query.get_or_404(form.product_id.data)
            quantity_to_remove = form.quantity.data
            
            # اعتبارسنجی
            if quantity_to_remove <= 0:
                flash("❌ تعداد باید بیشتر از صفر باشد", "danger")
                return render_template("inventory_write_off.html", form=form)
                
            if quantity_to_remove > product.quantity:
                flash(f"❌ تعداد درخواستی بیشتر از موجودی است! موجودی فعلی: {product.quantity} {product.unit}", "danger")
                return render_template("inventory_write_off.html", form=form)
            
            # محاسبه ضرر
            total_loss = quantity_to_remove * product.buying_price
            
            # کاهش موجودی محصول
            product.quantity -= quantity_to_remove
            
            # ثبت ضرر در دیتابیس
            inventory_loss = InventoryLoss(
                product_id=product.id,
                quantity=quantity_to_remove,
                unit_cost=product.buying_price,
                total_loss=total_loss,
                reason=form.reason.data,
                description=form.description.data,
                created_by=current_user.id
            )
            db.session.add(inventory_loss)
            db.session.commit()
            
            flash(f"✅ {quantity_to_remove} {product.unit} از {product.name} با موفقیت استاک‌اوت شد", "success")
            flash(f"📉 ضرر مالی ثبت شده: {total_loss:,.0f} افغانی", "info")
            return redirect(url_for('main_bp.inventory_loss_report'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ خطا در ثبت استاک‌اوت: {str(e)}", "error")
    
    return render_template("inventory_write_off.html", form=form)

# گزارش‌های موجودی
@main_bp.route('/inventory/loss-report')
@login_required
def inventory_loss_report():
    try:
        # گرفتن پارامترهای فیلتر از URL
        reason_filter = request.args.get('reason', 'all')
        product_filter = request.args.get('product_id', 'all')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # ایجاد کوئری پایه
        query = InventoryLoss.query
        
        # اعمال فیلترها
        if reason_filter != 'all':
            query = query.filter(InventoryLoss.reason == reason_filter)
        
        if product_filter != 'all':
            query = query.filter(InventoryLoss.product_id == product_filter)
        
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(InventoryLoss.loss_date >= start_date_obj)
        
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            query = query.filter(InventoryLoss.loss_date <= end_date_obj)
        
        # اجرای کوئری
        losses = query.order_by(InventoryLoss.loss_date.desc()).all()
        
        # محاسبه آمار
        total_loss = sum(loss.total_loss for loss in losses) if losses else 0
        total_quantity = sum(loss.quantity for loss in losses) if losses else 0
        
        # گرفتن لیست همه محصولات برای فیلتر
        all_products = Product.query.order_by(Product.name).all()
        
        return render_template('inventory_loss_report.html', 
                             losses=losses,
                             total_loss=total_loss,
                             total_quantity=total_quantity,
                             all_products=all_products,
                             reason_filter=reason_filter,
                             product_filter=product_filter,
                             start_date=start_date,
                             end_date=end_date)
                             
    except Exception as e:
        flash(f"خطا در بارگذاری گزارش ضرر: {str(e)}", "error")
        # مقادیر پیش‌فرض در صورت خطا
        return render_template('inventory_loss_report.html', 
                             losses=[],
                             total_loss=0,
                             total_quantity=0,
                             all_products=[],
                             reason_filter='all',
                             product_filter='all',
                             start_date='',
                             end_date='')
        
        
@main_bp.route('/inventory/expenses-report')
@login_required
def inventory_expenses_report():
    return render_template('expenses_report.html')

@main_bp.route("/creditors")
@login_required
def list_creditors():
    creditors = Creditor.query.order_by(Creditor.current_debt.desc()).all()
    
    # محاسبه مجموع بدهی‌ها
    total_debt = sum(c.current_debt for c in creditors)
    
    return render_template(
        "list_creditors.html",
        creditors=creditors,
        total_debt=total_debt
    )
    
@main_bp.route("/loans/add/<int:creditor_id>", methods=["GET", "POST"])
@login_required
def add_loan_for_creditor(creditor_id):
    """ثبت قرض جدید برای طلبکار مشخص"""
    creditor = Creditor.query.get_or_404(creditor_id)
    form = LoanForm()
    
    # تنظیم طلبکار پیش‌فرض
    form.creditor_id.choices = [(creditor.id, creditor.name)]
    form.creditor_id.data = creditor.id
    
    # دریافت لیست محصولات
    form.product_id.choices = [(0, 'بدون محصول')] + [
        (p.id, f"{p.name} - {p.batch_no or 'بدون بچ'}") 
        for p in Product.query.filter(Product.quantity > 0).all()
    ]
    
    if form.validate_on_submit():
        try:
            loan = Loan(
                creditor_id=creditor.id,
                product_id=form.product_id.data if form.product_id.data != 0 else None,
                amount=form.amount.data,
                paid_amount=0,
                quantity=form.quantity.data,
                unit=form.unit.data,
                date=form.date.data or datetime.now(),
                due_date=form.due_date.data,
                description=form.description.data,
                status='active'
            )
            
            db.session.add(loan)
            
            # به‌روزرسانی قرض داری جاری طلبکار
            creditor.current_debt += form.amount.data
            
            db.session.commit()
            
            flash(f"✅ قرض جدید برای {creditor.name} به مبلغ {form.amount.data:,.0f} افغانی ثبت شد", "success")
            return redirect(url_for('main_bp.list_creditors', open=creditor.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ خطا در ثبت قرض: {str(e)}", "error")
    
    return render_template("add_loan.html", form=form, creditor=creditor)
@main_bp.route("/creditors/edit/<int:creditor_id>", methods=["GET", "POST"])
@login_required
def edit_creditor(creditor_id):
    """ویرایش اطلاعات طلبکار"""
    creditor = Creditor.query.get_or_404(creditor_id)
    form = CreditorForm(obj=creditor)
    
    if form.validate_on_submit():
        try:
            # ذخیره قرض داری قبلی برای محاسبه تغییرات
            old_debt = creditor.current_debt
            
            # به‌روزرسانی اطلاعات
            creditor.name = form.name.data
            creditor.phone = form.phone.data
            creditor.address = form.address.data
            
            # اگر فیلد initial_debt در فرم وجود دارد
            if hasattr(form, 'initial_debt') and form.initial_debt.data is not None:
                creditor.initial_debt = form.initial_debt.data
            
            # اگر فیلد debt_description در فرم وجود دارد
            if hasattr(form, 'debt_description'):
                creditor.debt_description = form.debt_description.data
            
            # اگر فیلد current_debt در فرم وجود دارد و مقدار آن تغییر کرده
            if hasattr(form, 'current_debt') and form.current_debt.data is not None:
                creditor.current_debt = form.current_debt.data
            else:
                # اگر current_debt تغییر نکرده، همان مقدار قبلی را حفظ کن
                pass
            
            db.session.commit()
            
            flash(f"✅ اطلاعات طلبکار {creditor.name} با موفقیت ویرایش شد", "success")
            
            # بازگشت به صفحه طلبکاران با باز شدن جزئیات این طلبکار
            return redirect(url_for('main_bp.list_creditors', open=creditor.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ خطا در ویرایش طلبکار: {str(e)}", "error")
    
    return render_template("edit_creditor.html", form=form, creditor=creditor)
@main_bp.route("/creditors/pay/<int:creditor_id>", methods=["GET", "POST"])
@login_required
def pay_creditor(creditor_id):
    creditor = Creditor.query.get_or_404(creditor_id)
    
    # دریافت موجودی فعلی صندوق
    cash = CashBalance.query.first()
    if not cash:
        cash = CashBalance(amount=0, updated_by=current_user.id)
        db.session.add(cash)
        db.session.commit()
        print("✅ موجودی صندوق جدید ایجاد شد")
    
    print(f"💰 موجودی فعلی صندوق: {cash.amount:,.0f} افغانی")
    print(f"💰 بدهی جاری طلبکار {creditor.name}: {creditor.current_debt:,.0f} افغانی")
    
    if request.method == "POST":
        try:
            amount = float(request.form.get('amount', 0))
            receipt_number = request.form.get('receipt_number', '')
            description = request.form.get('description', '')
            payment_method = request.form.get('payment_method', 'نقدی')
            
            print(f"📥 درخواست پرداخت: {amount:,.0f} افغانی")
            
            # اعتبارسنجی
            if amount <= 0:
                flash("❌ مبلغ پرداخت باید مثبت باشد", "danger")
                return redirect(url_for("main_bp.pay_creditor", creditor_id=creditor.id))
            
            if amount > creditor.current_debt:
                flash(f"❌ مبلغ پرداختی ({amount:,.0f}) بیشتر از بدهی جاری ({creditor.current_debt:,.0f}) است", "danger")
                return redirect(url_for("main_bp.pay_creditor", creditor_id=creditor.id))
            
            # ✅ بررسی موجودی کافی صندوق (مهم)
            if cash.amount < amount:
                flash(f"❌ موجودی صندوق کافی نیست! موجودی فعلی: {cash.amount:,.0f} افغانی - مبلغ درخواستی: {amount:,.0f} افغانی", "danger")
                return redirect(url_for("main_bp.pay_creditor", creditor_id=creditor.id))
            
            # ========== 1. ثبت تراکنش پرداخت ==========
            transaction = DebtTransaction(
                creditor_id=creditor.id,
                user_id=current_user.id,
                amount=amount,
                transaction_type='payment',
                receipt_number=receipt_number,
                description=f"پرداخت به طلبکار {creditor.name} - {description}" if description else f"پرداخت به طلبکار {creditor.name}",
                date_created=datetime.utcnow()
            )
            db.session.add(transaction)
            print(f"✅ تراکنش پرداخت ثبت شد: ID {transaction.id}")
            
            # ========== 2. کاهش بدهی طلبکار ==========
            old_debt = creditor.current_debt
            creditor.current_debt -= amount
            print(f"✅ بدهی طلبکار از {old_debt:,.0f} به {creditor.current_debt:,.0f} کاهش یافت")
            
            # ========== 3. کاهش موجودی نقدی صندوق ✅ ==========
            old_cash_balance = cash.amount
            cash.amount -= amount
            cash.last_updated = datetime.utcnow()
            cash.updated_by = current_user.id
            print(f"✅ موجودی صندوق از {old_cash_balance:,.0f} به {cash.amount:,.0f} کاهش یافت")
            
            # ========== 4. ثبت تراکنش نقدی ==========
            cash_transaction = CashTransaction(
                amount=-amount,
                transaction_type='creditor_payment',
                description=f"پرداخت به طلبکار {creditor.name} - مبلغ: {amount:,.0f} افغانی - روش: {payment_method} - {receipt_number or 'بدون رسید'}",
                balance_before=old_cash_balance,
                balance_after=cash.amount,
                created_by=current_user.id,
                reference_id=transaction.id
            )
            db.session.add(cash_transaction)
            print(f"✅ تراکنش نقدی ثبت شد: {cash_transaction.id}")
            
            # به‌روزرسانی تاریخ آخرین تراکنش طلبکار
            creditor.last_transaction_date = datetime.utcnow()
            creditor.last_transaction_description = f"پرداخت مبلغ {amount:,.0f} افغانی - موجودی باقیمانده: {creditor.current_debt:,.0f}"
            
            db.session.commit()
            
            flash(f"✅ پرداخت به مبلغ {amount:,.0f} افغانی به {creditor.name} ثبت شد", "success")
            flash(f"💰 موجودی قبلی صندوق: {old_cash_balance:,.0f} افغانی", "info")
            flash(f"💰 موجودی فعلی صندوق: {cash.amount:,.0f} افغانی", "info")
            
            return redirect(url_for("main_bp.list_creditors"))
            
        except Exception as e:
            db.session.rollback()
            print(f"❌ خطا: {str(e)}")
            import traceback
            traceback.print_exc()
            flash(f"❌ خطا در ثبت پرداخت: {str(e)}", "error")
    
    return render_template("pay_creditor.html", creditor=creditor, cash_balance=cash.amount)

@main_bp.route("/creditors/add", methods=["GET", "POST"])
@login_required
def add_creditor():
    form = CreditorForm()
    
    if form.validate_on_submit():
        try:
            creditor = Creditor(
                name=form.name.data,
                phone=form.phone.data,
                address=form.address.data,
                initial_debt=form.initial_debt.data,
                current_debt=form.initial_debt.data,
                debt_description=form.debt_description.data,  # 🔥 اضافه شده
                created_by=current_user.id
            )
            
            db.session.add(creditor)
            db.session.commit()
            
            flash("طلبکار با موفقیت ثبت شد", "success")
            return redirect(url_for('main_bp.list_creditors'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"خطا در ثبت طلبکار: {str(e)}", "error")
    
    return render_template("add_creditor.html", form=form)


@main_bp.route('/debt-report')
@login_required
def api_creditors_debt():
    creditors = Creditor.query.all()
    total_debt = sum(c.current_debt for c in creditors)
    return render_template("debt_report.html", 
                         creditors=creditors, 
                         total_debt=total_debt)

@main_bp.route("/products/edit/<int:product_id>", methods=["GET", "POST"])
@login_required
def edit_product(product_id):
    product = Product.query.get_or_404(product_id)
    form = ProductForm(obj=product)

    # تنظیم choices طلبکاران
    try:
        creditors = Creditor.query.order_by(Creditor.name).all()
        form.creditor_id.choices = [(0, "انتخاب طلبکار")] + [
            (c.id, f"{c.name} (قرض داری: {c.current_debt:,.0f} افغانی)") for c in creditors
        ]
    except Exception as e:
        form.creditor_id.choices = [(0, "انتخاب طلبکار")]

    if request.method == 'GET':
        form.creditor_id.data = product.creditor_id if product.creditor_id else 0
        form.purchase_type.data = product.purchase_type
        return render_template("edit_product.html", form=form, product=product, creditors=creditors)

    # ========== POST - نادیده گرفتن اعتبارسنجی فرم ==========
    try:
        # ذخیره اطلاعات قدیمی
        old_creditor_id = product.creditor_id
        old_purchase_type = product.purchase_type
        old_buying_price = product.buying_price
        old_quantity = product.quantity
        old_items_per_carton = product.items_per_carton
        old_single_quantity = product.single_quantity
        
        # محاسبه تعداد کل قدیمی به دانه
        old_total_items = (old_quantity * old_items_per_carton) + old_single_quantity
        old_total_cost = old_buying_price * old_total_items
        
        # ======================
        # دریافت اطلاعات جدید از فرم
        # ======================
        
        # اطلاعات پایه
        product.name = request.form.get('name')
        product.batch_no = request.form.get('batch_no')
        
        # بارکد
        new_barcode = request.form.get('barcode', '').strip() or None
        if new_barcode and new_barcode != product.barcode:
            existing = Product.query.filter_by(barcode=new_barcode).first()
            if existing and existing.id != product.id:
                flash(f"❌ بارکد {new_barcode} قبلاً ثبت شده است", "danger")
                return render_template("edit_product.html", form=form, product=product, creditors=creditors)
        product.barcode = new_barcode
        
        # ✅ موجودی جدید
        carton_quantity = float(request.form.get('carton_quantity', 0) or 0)
        items_per_carton = float(request.form.get('items_per_carton', 1) or 1)
        single_quantity = float(request.form.get('single_quantity', 0) or 0)
        
        product.quantity = carton_quantity
        product.items_per_carton = items_per_carton
        product.single_quantity = single_quantity
        
        # محاسبه تعداد کل جدید به دانه
        new_total_items = (carton_quantity * items_per_carton) + single_quantity
        
        # قیمت‌ها
        product.buying_price = float(request.form.get('buying_price', 0) or 0)
        product.selling_price = float(request.form.get('selling_price', 0) or 0)
        
        # محاسبه هزینه کل جدید
        new_total_cost = product.buying_price * new_total_items
        
        # واحد
        product.unit = request.form.get('unit', 'عدد')
        
        # تاریخ انقضا
        expiry_date_str = request.form.get('expiry_date')
        if expiry_date_str:
            try:
                product.expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
            except:
                product.expiry_date = None
        else:
            product.expiry_date = None
        
        # نوع خرید و طلبکار
        product.purchase_type = request.form.get('purchase_type', 'cash')
        creditor_id = request.form.get('creditor_id', 0)
        product.creditor_id = int(creditor_id) if creditor_id and int(creditor_id) != 0 else None
        
        # توضیحات
        product.purchase_description = request.form.get('purchase_description', '')
        
        # ======================
        # مدیریت قرض داری طلبکار (مهم)
        # ======================
        
        # 1. اگر قبلاً قرضی بوده و تغییر کرده است - حذف قرض قبلی
        if old_purchase_type == 'credit' and old_creditor_id:
            old_creditor = Creditor.query.get(old_creditor_id)
            if old_creditor and old_total_cost > 0:
                old_creditor.current_debt -= old_total_cost
                if old_creditor.current_debt < 0:
                    old_creditor.current_debt = 0
                
                # ثبت تراکنش برای حذف قرض قبلی
                transaction = DebtTransaction(
                    creditor_id=old_creditor.id,
                    amount=old_total_cost,
                    transaction_type='payment',
                    description=f"ویرایش محصول: حذف قرض قبلی - {product.name} - تعداد: {int(old_total_items)} عدد - قیمت: {old_buying_price:,.0f} افغانی",
                    receipt_number=f"EDIT-{product.id}-OLD",
                    date_created=datetime.now(),
                    user_id=current_user.id
                )
                db.session.add(transaction)
                flash(f"✅ قرض قبلی {old_creditor.name} به مقدار {old_total_cost:,.0f} افغانی کاهش یافت", "info")
        
        # 2. اگر خرید جدید به صورت قرضی است - اضافه کردن قرض جدید
        if product.purchase_type == 'credit' and product.creditor_id and new_total_cost > 0:
            new_creditor = Creditor.query.get(product.creditor_id)
            if new_creditor:
                new_creditor.current_debt += new_total_cost
                
                # ثبت تراکنش برای قرض جدید
                transaction_description = f"ویرایش محصول: {product.name} - تعداد کارتن: {carton_quantity} - تعداد در کارتن: {items_per_carton} - تعداد تکی: {single_quantity} - مجموع کل: {int(new_total_items)} عدد - قیمت خرید: {product.buying_price:,.0f} افغانی"
                
                if product.batch_no:
                    transaction_description += f" - بچ: {product.batch_no}"
                if product.expiry_date:
                    transaction_description += f" - انقضا: {product.expiry_date.strftime('%Y-%m-%d')}"
                if product.purchase_description:
                    transaction_description += f" - توضیحات: {product.purchase_description}"
                
                transaction = DebtTransaction(
                    creditor_id=new_creditor.id,
                    amount=new_total_cost,
                    transaction_type='debt',
                    description=transaction_description,
                    receipt_number=f"EDIT-{product.id}-NEW",
                    date_created=datetime.now(),
                    user_id=current_user.id
                )
                db.session.add(transaction)
                flash(f"✅ قرض جدید {new_creditor.name} به مقدار {new_total_cost:,.0f} افغانی افزایش یافت", "info")
        
        # 3. اگر از قرضی به نقدی تغییر کرده است - فقط حذف قرض قبلی (قبلاً انجام شد)
        elif old_purchase_type == 'credit' and product.purchase_type != 'credit':
            # قرض قبلی در مرحله 1 حذف شده است
            flash("✅ خرید از قرضی به نقدی تغییر کرد.", "info")
        
        # 4. اگر از نقدی به قرضی تغییر کرده است - اضافه کردن قرض جدید
        elif old_purchase_type != 'credit' and product.purchase_type == 'credit' and product.creditor_id:
            # قرض جدید در مرحله 2 اضافه شده است
            flash("✅ خرید از نقدی به قرضی تغییر کرد.", "info")
        
        # 5. اگر قرضی بوده و طلبکار تغییر کرده است
        elif old_purchase_type == 'credit' and product.purchase_type == 'credit' and old_creditor_id != product.creditor_id:
            # قرض قبلی در مرحله 1 حذف شد
            # قرض جدید در مرحله 2 اضافه شد
            flash(f"✅ طلبکار از {old_creditor.name if old_creditor else 'نامشخص'} به {new_creditor.name if new_creditor else 'جدید'} تغییر کرد.", "info")
        
        # 6. اگر قرضی بوده و تعداد یا قیمت تغییر کرده است
        elif old_purchase_type == 'credit' and product.purchase_type == 'credit' and old_creditor_id == product.creditor_id:
            if old_total_cost != new_total_cost:
                # تفاوت قیمت را محاسبه کن
                diff_cost = new_total_cost - old_total_cost
                if diff_cost != 0:
                    creditor = Creditor.query.get(product.creditor_id)
                    if creditor:
                        creditor.current_debt += diff_cost
                        
                        # ثبت تراکنش برای تفاوت قیمت
                        transaction_type = 'debt' if diff_cost > 0 else 'payment'
                        transaction = DebtTransaction(
                            creditor_id=creditor.id,
                            amount=abs(diff_cost),
                            transaction_type=transaction_type,
                            description=f"ویرایش محصول: تغییر تعداد/قیمت - {product.name} - تغییر: {diff_cost:,.0f} افغانی",
                            receipt_number=f"EDIT-{product.id}-DIFF",
                            date_created=datetime.now(),
                            user_id=current_user.id
                        )
                        db.session.add(transaction)
                        
                        if diff_cost > 0:
                            flash(f"✅ قرض داری به {creditor.name} به مقدار {diff_cost:,.0f} افغانی افزایش یافت", "info")
                        else:
                            flash(f"✅ قرض داری به {creditor.name} به مقدار {abs(diff_cost):,.0f} افغانی کاهش یافت", "info")
        
        # ذخیره نهایی
        db.session.commit()
        
        print(f"✅ محصول {product.name} با موفقیت ویرایش شد")
        print(f"   هزینه قدیم: {old_total_cost:,.0f}, هزینه جدید: {new_total_cost:,.0f}")
        print(f"   نوع خرید قدیم: {old_purchase_type}, نوع خرید جدید: {product.purchase_type}")
        
        flash("✅ محصول با موفقیت ویرایش شد", "success")
        return redirect(url_for('main_bp.list_products'))
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ خطا در ویرایش محصول: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f"❌ خطا در ویرایش محصول: {str(e)}", "danger")
        return render_template("edit_product.html", form=form, product=product, creditors=creditors)
            
@main_bp.route("/products/delete/<int:product_id>", methods=["POST"])
@login_required
def delete_product(product_id):
    product = Product.query.get_or_404(product_id)
    try:
        db.session.delete(product)
        db.session.commit()
        flash("✅ محصول با موفقیت حذف شد", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ خطا در حذف محصول: {str(e)}", "error")
        print(f"❌ خطا در حذف محصول: {str(e)}")
    return redirect(url_for('main_bp.list_products'))

# ==================== سیستم تراکنش طلبکاران ====================

# 1. اول route تست را تعریف کنید
@main_bp.route("/debug-creditors")
def debug_creditors():
    """تست routeها"""
    from flask import jsonify, url_for
    return jsonify({
        "message": "Debug route works",
        "blueprint_name": main_bp.name,
        "endpoint": "main_bp.creditor_transactions",
        "url_for_test": str(url_for("main_bp.creditor_transactions", creditor_id=1))
    })

# در بالای routes.py، این import را اضافه کنید

@main_bp.route("/creditors/transactions", methods=["GET", "POST"])
@main_bp.route("/creditors/<int:creditor_id>/transactions", methods=["GET", "POST"])
@login_required
def creditor_transactions(creditor_id=None):
    """مدیریت تراکنش‌های طلبکار با به‌روزرسانی موجودی نقدی"""
    from app.models import CashBalance, CashTransaction
    
    # اگر ID داده نشده، اولین طلبکار را بگیر
    if creditor_id is None:
        first_creditor = Creditor.query.first()
        if first_creditor:
            return redirect(url_for('main_bp.creditor_transactions', creditor_id=first_creditor.id))
        else:
            flash("❗ هیچ طلبکاری ثبت نشده است. لطفاً ابتدا طلبکار ثبت کنید.", "warning")
            return redirect(url_for('main_bp.index'))
    
    creditor = Creditor.query.get(creditor_id)
    if not creditor:
        flash(f"❌ طلبکاری با شناسه {creditor_id} یافت نشد.", "danger")
        return redirect(url_for('main_bp.index'))
    
    # دریافت موجودی فعلی صندوق
    cash = CashBalance.query.first()
    if not cash:
        cash = CashBalance(amount=0)
        db.session.add(cash)
        db.session.commit()
    
    form = TransactionForm()
    all_creditors = Creditor.query.order_by(Creditor.name).all()
    form.transaction_date.data = datetime.utcnow().date()

    # ثبت تراکنش جدید
    if form.validate_on_submit():
        try:
            amount = float(form.amount.data)
            transaction_type = form.transaction_type.data
            description = form.description.data
            receipt_number = form.receipt_number.data
            transaction_date = form.transaction_date.data or datetime.utcnow().date()
            
            old_debt = creditor.current_debt
            
            # ========== پرداخت به طلبکار (کاهش بدهی) ==========
            if transaction_type == 'payment':
                # ✅ بررسی موجودی کافی صندوق
                if cash.amount < amount:
                    flash(f"❌ موجودی صندوق کافی نیست! موجودی فعلی: {cash.amount:,.0f} افغانی - مبلغ درخواستی: {amount:,.0f} افغانی", "danger")
                    return redirect(url_for("main_bp.creditor_transactions", creditor_id=creditor.id))
                
                if amount > creditor.current_debt:
                    flash(f"⚠️ مبلغ پرداختی ({amount:,.0f}) بیشتر از قرض ({creditor.current_debt:,.0f}) است", "warning")
                    return redirect(url_for("main_bp.creditor_transactions", creditor_id=creditor.id))
                
                # کاهش بدهی طلبکار
                creditor.current_debt -= amount
                
                # ✅ کاهش موجودی نقدی صندوق
                old_cash_balance = cash.amount
                cash.amount -= amount
                cash.last_updated = datetime.utcnow()
                cash.updated_by = current_user.id
                
                # ثبت تراکنش نقدی
                cash_transaction = CashTransaction(
                    amount=-amount,
                    transaction_type='creditor_payment',
                    description=f"پرداخت به طلبکار {creditor.name} - {description or 'بدون توضیحات'}",
                    balance_before=old_cash_balance,
                    balance_after=cash.amount,
                    created_by=current_user.id
                )
                db.session.add(cash_transaction)
                
                flash_message = f"✅ پرداخت به مبلغ {amount:,.0f} افغانی به {creditor.name} ثبت شد. قرض از {old_debt:,.0f} به {creditor.current_debt:,.0f} افغانی کاهش یافت"
                flash(f"💰 موجودی فعلی صندوق: {cash.amount:,.0f} افغانی", "info")
                
            # ========== ثبت قرض جدید (افزایش بدهی) ==========
            else:  # transaction_type == 'debt'
                creditor.current_debt += amount
                flash_message = f"✅ قرض جدید به مبلغ {amount:,.0f} افغانی ثبت شد. قرض از {old_debt:,.0f} به {creditor.current_debt:,.0f} افغانی افزایش یافت"
            
            # ثبت تراکنش DebtTransaction
            transaction = DebtTransaction(
                creditor_id=creditor.id,
                amount=amount,
                transaction_type=transaction_type,
                description=description,
                receipt_number=receipt_number,
                date_created=datetime.combine(transaction_date, datetime.min.time()),
                user_id=current_user.id
            )
            
            db.session.add(transaction)
            db.session.commit()
            flash(flash_message, "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ خطا در ثبت تراکنش: {str(e)}", "error")
        
        return redirect(url_for("main_bp.creditor_transactions", creditor_id=creditor.id))

    # تاریخچه تراکنش‌ها
    transactions = DebtTransaction.query.filter_by(creditor_id=creditor.id)\
                    .order_by(DebtTransaction.date_created.desc()).all()
    
    return render_template(
        "creditor_transactions.html",
        creditor=creditor,
        form=form,
        transactions=transactions,
        all_creditors=all_creditors,
        cash_balance=cash.amount
    )
# ==================== خروجی Excel ====================
@main_bp.route("/export/products")
@login_required
def export_products():
    """خروجی Excel از اجناس با سیستم موجودی دوگانه (کارتن + تکی)"""

    products = Product.query.order_by(Product.name).all()

    data = []

    for product in products:

        # ======================
        # اطلاعات پایه
        # ======================
        
        # تعداد کارتن در انبار
        carton_quantity = product.quantity or 0
        
        # تعداد در هر کارتن
        items_per_carton = product.items_per_carton or 1
        
        # تعداد تکی
        single_quantity = product.single_quantity or 0
        
        # ======================
        # محاسبه مجموع دانه‌ها
        # فرمول: (تعداد کارتن × تعداد در کارتن) + تعداد تکی
        # ======================
        
        total_items = (carton_quantity * items_per_carton) + single_quantity
        
        # ======================
        # محاسبه کارتن‌های کامل و تکی باقیمانده از مجموع کل
        # (برای نمایش دقیق)
        # ======================
        
        full_cartons = int(total_items // items_per_carton) if items_per_carton > 0 else 0
        remaining_singles = int(total_items % items_per_carton) if items_per_carton > 0 else int(total_items)
        
        # ======================
        # ارزش موجودی
        # ======================
        
        stock_value = total_items * float(product.selling_price or 0)
        
        # ======================
        # وضعیت موجودی (بر اساس مجموع کل)
        # ======================
        
        stock_status = "عادی"
        min_stock = product.min_stock or 0
        
        if total_items <= 0:
            stock_status = "تمام شده"
        elif total_items <= min_stock:
            stock_status = "موجودی کم"
        
        # ======================
        # وضعیت انقضا
        # ======================
        
        expiry_status = "---"
        
        if product.expiry_date:
            remaining_days = (product.expiry_date - datetime.now().date()).days
            
            if remaining_days < 0:
                expiry_status = "منقضی شده"
            elif remaining_days <= 30:
                expiry_status = f"نزدیک انقضا ({remaining_days} روز)"
            else:
                expiry_status = "سالم"

        data.append({
            # ======================
            # اطلاعات اصلی
            # ======================
            "شناسه": product.id,
            "نام جنس": product.name,
            "بارکد": product.barcode or "---",
            
            # ======================
            # موجودی به صورت خام
            # ======================
            "تعداد کارتن (موجودی)": carton_quantity,
            "تعداد در هر کارتن": items_per_carton,
            "تعداد تکی (موجودی)": single_quantity,
            
            # ======================
            # موجودی محاسبه شده
            # ======================
            "مجموع دانه (کل)": total_items,
            "کارتن کامل از کل": full_cartons,
            "تکی باقیمانده از کل": remaining_singles,
            
            # ======================
            # قیمت‌ها
            # ======================
            "قیمت خرید یک دانه": format_currency(product.buying_price),
            "قیمت فروش یک دانه": format_currency(product.selling_price),
            
            # ======================
            # ارزش مالی
            # ======================
            "ارزش موجودی (فروش)": format_currency(stock_value),
            
            # ======================
            # وضعیت موجودی
            # ======================
            "حداقل موجودی (اخطار)": min_stock,
            "وضعیت موجودی": stock_status,
            
            # ======================
            # دسته‌بندی
            # ======================
            "دسته‌بندی": product.category or "عمومی",
            "واحد": product.unit or "عدد",
            
            # ======================
            # انقضا
            # ======================
            "شماره بچ (Batch)": product.batch_no or "---",
            "تاریخ انقضا": product.expiry_date.strftime('%Y-%m-%d') if product.expiry_date else "---",
            "وضعیت انقضا": expiry_status,
            
            # ======================
            # خرید
            # ======================
            "نوع خرید": "نقدی" if product.purchase_type == "cash" else "قرضی",
            "طلبکار": product.creditor.name if product.creditor else "---",
            
            # ======================
            # توضیحات
            # ======================
            "توضیحات خرید": product.purchase_description or "---",
            
            # ======================
            # تاریخ‌ها
            # ======================
            "تاریخ ثبت": product.created_at.strftime('%Y-%m-%d %H:%M') if product.created_at else "---",
            "آخرین بروزرسانی": product.updated_at.strftime('%Y-%m-%d %H:%M') if product.updated_at else "---"
        })

    # اگر داده خالی بود
    if not data:
        data.append({
            "پیام": "هیچ محصولی در سیستم ثبت نشده است"
        })

    return export_to_excel(
        data,
        f"products_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "اجناس",
        "گزارش کامل اجناس فروشگاه (سیستم دوگانه کارتن و تکی)"
    )
@main_bp.route("/export")
@login_required
def export_options():
    """صفحه انتخاب نوع خروجی"""
    from app.models import Customer, Creditor, Sale, Product, DailyExpense
    from app import db
    
    debtors_count = Customer.query.filter(Customer.total_debt > 0).count()
    creditors_count = Creditor.query.filter(Creditor.current_debt > 0).count()
    sales_count = Sale.query.count()
    products_count = Product.query.count()
    customers_count = Customer.query.count()
    expenses_count = DailyExpense.query.count()
    
    # آمار مالی
    total_sales = db.session.query(db.func.sum(Sale.final_amount)).scalar() or 0
    total_expenses = db.session.query(db.func.sum(DailyExpense.amount)).scalar() or 0
    total_customer_debt = db.session.query(db.func.sum(Customer.total_debt)).scalar() or 0
    total_creditor_debt = db.session.query(db.func.sum(Creditor.current_debt)).scalar() or 0
    
    return render_template(
        "export_options.html",
        debtors_count=debtors_count,
        creditors_count=creditors_count,
        sales_count=sales_count,
        products_count=products_count,
        customers_count=customers_count,
        expenses_count=expenses_count,
        total_sales=total_sales,
        total_expenses=total_expenses,
        total_customer_debt=total_customer_debt,
        total_creditor_debt=total_creditor_debt
    )
    
@main_bp.route("/export/debtors")
@login_required
def export_debtors():
    """خروجی Excel از بدهکاران (مشتریان بدهکار)"""
    # مشتریانی که قرض داری دارند
    debtors = Customer.query.filter(Customer.total_debt > 0).order_by(Customer.total_debt.desc()).all()
    
    data = []
    for debtor in debtors:
        # محاسبه تعداد فروش‌های نسیه این مشتری
        credit_sales = Sale.query.filter(
            Sale.customer_id == debtor.id,
            Sale.remaining_debt > 0
        ).count()
        
        data.append({
            "شناسه": debtor.id,
            "نام مشتری": debtor.name,
            "شماره تماس": debtor.phone or "---",
            "آدرس": debtor.address or "---",
            "مجموع قرض داری": format_currency(debtor.total_debt),
            "تعداد فروش نسیه": credit_sales,
            "تاریخ ثبت": debtor.created_at.strftime('%Y-%m-%d') if debtor.created_at else "---"
        })
    
    return export_to_excel(
        data, 
        f"debtors_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "بدهکاران",
        "لیست بدهکاران (مشتریان با قرض داری)"
    )
@main_bp.route("/export/creditors")
@login_required
def export_creditors():
    """خروجی Excel از طلبکاران"""
    # ✅ حذف شرط current_debt > 0
    creditors = Creditor.query.order_by(Creditor.current_debt.desc()).all()
    
    data = []
    for creditor in creditors:
        products_count = Product.query.filter_by(creditor_id=creditor.id).count()
        
        data.append({
            "شناسه": creditor.id,
            "نام طلبکار": creditor.name,
            "شماره تماس": creditor.phone or "---",
            "آدرس": creditor.address or "---",
            "قرض داری اولیه": format_currency(creditor.initial_debt),
            "قرض داری جاری": format_currency(creditor.current_debt),
            "توضیحات": creditor.debt_description or "---",
            "تعداد محصولات": products_count,
            "تاریخ ثبت": creditor.created_at.strftime('%Y-%m-%d') if creditor.created_at else "---"
        })
    
    # ✅ پیام اگر داده‌ای وجود نداشت
    if not data:
        data.append({"پیام": "هیچ طلبکاری در سیستم ثبت نشده است"})
    
    return export_to_excel(
        data,
        f"creditors_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "طلبکاران",
        "لیست طلبکاران"
    )
@main_bp.route("/export/sales")
@login_required
def export_sales():
    sales = Sale.query.order_by(Sale.date.desc()).all()
    
    data = []
    for sale in sales:
        total_profit = sum(item.profit or 0 for item in sale.items)  # ✅ استفاده از profit ذخیره شده
        
        data.append({
            "شناسه": sale.id,
            "شماره فاکتور": sale.invoice_number,
            "تاریخ": sale.date.strftime('%Y-%m-%d'),
            "مشتری": sale.customer.name if sale.customer else "عمومی",
            "مبلغ کل": format_currency(sale.total_amount),
            "تخفیف": format_currency(sale.total_discount),
            "مبلغ نهایی": format_currency(sale.final_amount),
            "پرداختی": format_currency(sale.amount_paid),
            "باقی‌مانده": format_currency(sale.remaining_debt),
            "مفاد خالص": format_currency(total_profit),  # ✅ مفاد واقعی
            "تعداد اقلام": len(sale.items),
        })
    
    return export_to_excel(data, f"sales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", "فروشات", "گزارش فروشات")

@main_bp.route("/export/customers")
@login_required
def export_customers():
    """خروجی Excel از مشتری‌ها"""
    customers = Customer.query.order_by(Customer.name).all()
    
    data = []
    for customer in customers:
        # محاسبه تعداد خریدها و مجموع خرید
        sales_count = Sale.query.filter_by(customer_id=customer.id).count()
        total_purchases = db.session.query(db.func.sum(Sale.final_amount)).filter(
            Sale.customer_id == customer.id
        ).scalar() or 0
        
        data.append({
            "شناسه": customer.id,
            "نام مشتری": customer.name,
            "شماره تماس": customer.phone or "---",
            "آدرس": customer.address or "---",
            "قرض داری جاری": format_currency(customer.total_debt or 0),
            "تعداد خرید": sales_count,
            "مجموع خرید": format_currency(total_purchases),
            "تاریخ ثبت": customer.created_at.strftime('%Y-%m-%d') if customer.created_at else "---",
            "آخرین بروزرسانی": customer.updated_at.strftime('%Y-%m-%d') if customer.updated_at else "---"
        })
    
    return export_to_excel(
        data,
        f"customers_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "مشتریان",
        "لیست تمام مشتریان"
    )

@main_bp.route("/export/low_stock")
@login_required
def export_low_stock():
    """خروجی Excel از اجناس با موجودی کم"""
    products = Product.query.filter(
        Product.quantity <= Product.min_stock
    ).order_by(Product.quantity).all()
    
    data = []
    for product in products:
        shortage = product.min_stock - product.quantity if product.min_stock else 0
        
        data.append({
            "شناسه": product.id,
            "نام جنس": product.name,
            "موجودی فعلی": f"{product.quantity} {product.unit}",
            "حداقل موجودی": product.min_stock or 0,
            "کمبود": f"{shortage} {product.unit}" if shortage > 0 else "---",
            "قیمت فروش": format_currency(product.selling_price),
            "دسته‌بندی": product.category or "عمومی",
            "تاریخ انقضا": product.expiry_date.strftime('%Y-%m-%d') if product.expiry_date else "---"
        })
    
    return export_to_excel(
        data,
        f"low_stock_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "موجودی کم",
        "اجناس با موجودی کمتر از حد مجاز"
    )

@main_bp.route("/export/expiring")
@login_required
def export_expiring():
    """خروجی Excel از اجناس نزدیک به انقضا"""
    from datetime import datetime, timedelta
    
    today = datetime.now().date()
    thirty_days_later = today + timedelta(days=30)
    
    products = Product.query.filter(
        Product.expiry_date >= today,
        Product.expiry_date <= thirty_days_later
    ).order_by(Product.expiry_date).all()
    
    data = []
    for product in products:
        remaining_days = (product.expiry_date - today).days
        
        data.append({
            "شناسه": product.id,
            "نام جنس": product.name,
            "موجودی": f"{product.quantity} {product.unit}",
            "تاریخ انقضا": product.expiry_date.strftime('%Y-%m-%d'),
            "روزهای باقی‌مانده": remaining_days,
            "قیمت فروش": format_currency(product.selling_price),
            "دسته‌بندی": product.category or "عمومی"
        })
    
    return export_to_excel(
        data,
        f"expiring_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "نزدیک انقضا",
        "اجناس نزدیک به تاریخ انقضا (۳۰ روز)"
    )
    
@main_bp.route("/cash/balance")
@login_required
def cash_balance():
    """نمایش موجودی نقدی و تاریخچه تراکنش‌ها"""
    from app.models import CashBalance, CashTransaction
    
    cash = CashBalance.query.first()
    if not cash:
        cash = CashBalance(amount=0, updated_by=current_user.id)
        db.session.add(cash)
        db.session.commit()
    
    # ✅ تاریخچه تراکنش‌ها
    transactions = CashTransaction.query.order_by(CashTransaction.created_at.desc()).limit(100).all()
    
    # محاسبه آمار
    total_sales = sum(t.amount for t in transactions if t.transaction_type == 'sale' and t.amount > 0)
    total_expenses = sum(abs(t.amount) for t in transactions if t.transaction_type in ['expense', 'withdrawal'] and t.amount < 0)
    
    return render_template(
        "cash_balance.html",
        cash=cash,
        transactions=transactions,
        total_sales=total_sales,
        total_expenses=total_expenses
    )
    
@main_bp.route("/cash/update", methods=["POST"])
@login_required
def cash_update():
    """به‌روزرسانی موجودی نقدی"""
    try:
        amount = float(request.form.get('amount', 0))
        description = request.form.get('description', '')
        
        cash = CashBalance.query.first()
        if not cash:
            cash = CashBalance(amount=0)
            db.session.add(cash)
        
        old_balance = cash.amount
        cash.amount = amount
        cash.updated_by = current_user.id
        
        transaction = CashTransaction(
            amount=amount - old_balance,
            transaction_type='adjustment',
            description=description or "تنظیم دستی موجودی",
            balance_before=old_balance,
            balance_after=amount,
            created_by=current_user.id
        )
        db.session.add(transaction)
        db.session.commit()
        
        flash(f"✅ موجودی نقدی به {amount:,.0f} افغانی تنظیم شد", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"❌ خطا: {str(e)}", "error")
    
    # ✅ این خط را چک کنید - باید به مسیر /cash/balance برود
    return redirect(url_for('main_bp.cash_balance'))  # این درست است چون cash_balance برای /cash/balance است
@main_bp.route("/dashboard")
@login_required
def dashboard():
    """داشبورد اصلی"""
    from datetime import datetime, timedelta
    from app.models import Sale, Expense
    
    # ===== داده‌های نمودار فروش روزانه (۷ روز اخیر) =====
    today = datetime.now().date()
    daily_sales_labels = []
    daily_sales_data = []
    
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        daily_sales_labels.append(day.strftime('%Y-%m-%d'))
        
        # محاسبه فروش آن روز
        start_of_day = datetime.combine(day, datetime.min.time())
        end_of_day = datetime.combine(day, datetime.max.time())
        
        daily_total = db.session.query(db.func.sum(Sale.final_amount)).filter(
            Sale.date >= start_of_day,
            Sale.date <= end_of_day
        ).scalar() or 0
        
        daily_sales_data.append(float(daily_total))
    
    # ===== داده‌های نمودار هزینه‌ها =====
    # گروه‌بندی هزینه‌ها بر اساس دسته‌بندی
    expenses_by_category = db.session.query(
        Expense.category,
        db.func.sum(Expense.amount)
    ).group_by(Expense.category).all()
    
    expenses_labels = [item[0] for item in expenses_by_category]
    expenses_data = [float(item[1]) for item in expenses_by_category]
    
    return render_template(
        "dashboard.html",
        daily_sales_labels=daily_sales_labels,
        daily_sales_data=daily_sales_data,
        expenses_labels=expenses_labels,
        expenses_data=expenses_data
    )
    
@main_bp.route("/cash/withdraw", methods=["GET", "POST"])
@login_required
def cash_withdraw():
    """برداشت وجه از صندوق توسط کارمندان"""
    from app.models import CashWithdrawal
    from datetime import datetime
    
    if request.method == "POST":
        try:
            amount = float(request.form.get('amount', 0))
            purpose = request.form.get('purpose', '').strip()
            description = request.form.get('description', '').strip()
            
            if amount <= 0:
                flash("❌ مبلغ باید بزرگتر از صفر باشد", "error")
                return redirect(url_for('main_bp.cash_withdraw'))
            
            if not purpose:
                flash("❌ لطفاً هدف برداشت را مشخص کنید", "error")
                return redirect(url_for('main_bp.cash_withdraw'))
            
            # بررسی موجودی کافی
            cash = CashBalance.query.first()
            if not cash or cash.amount < amount:
                flash(f"❌ موجودی نقدی کافی نیست. موجودی: {cash.amount if cash else 0:,.0f} افغانی", "error")
                return redirect(url_for('main_bp.cash_withdraw'))
            
            # ثبت برداشت
            withdrawal = CashWithdrawal(
                employee_id=current_user.id,
                amount=amount,
                purpose=purpose,
                description=description,  # ✅ اینجا توضیحات ذخیره می‌شود
                withdrawal_date=datetime.now()
            )
            db.session.add(withdrawal)
            db.session.flush()
            
            # کاهش موجودی نقدی
            success = update_cash_balance(
                amount=-amount,
                transaction_type='withdrawal',
                reference_id=withdrawal.id,
                description=f"برداشت توسط {current_user.full_name} - {purpose} - {description}"  # ✅ توضیحات در تراکنش
            )
            
            if success:
                db.session.commit()
                flash(f"✅ برداشت {amount:,.0f} افغانی توسط {current_user.full_name} ثبت شد", "success")
                if description:
                    flash(f"📝 توضیحات: {description}", "info")
                flash(f"💰 هدف: {purpose}", "info")
            else:
                db.session.rollback()
                flash("❌ خطا در ثبت برداشت", "error")
            
            return redirect(url_for('main_bp.cash_balance'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ خطا: {str(e)}", "error")
            return redirect(url_for('main_bp.cash_withdraw'))
    
    # نمایش فرم برداشت
    cash = CashBalance.query.first()
    return render_template("cash_withdraw.html", cash=cash, now=datetime.now())

@main_bp.route("/cash/initial", methods=["GET", "POST"])
@login_required
def cash_initial():
    """ثبت موجودی اولیه صندوق (فقط برای مدیر)"""
    from datetime import datetime
    
    # فقط مدیر می‌تواند موجودی اولیه را ثبت کند
    if current_user.role != 'manager':
        flash("❌ فقط مدیر می‌تواند موجودی اولیه را ثبت کند", "error")
        return redirect(url_for('main_bp.index'))
    
    # بررسی آیا قبلاً موجودی ثبت شده
    cash = CashBalance.query.first()
    if cash and cash.amount != 0:
        flash(f"⚠️ قبلاً موجودی اولیه ثبت شده است. موجودی فعلی: {cash.amount:,.0f} افغانی", "warning")
        return redirect(url_for('main_bp.cash_balance'))
    
    if request.method == "POST":
        try:
            amount = float(request.form.get('amount', 0))
            description = request.form.get('description', '').strip()
            
            if amount <= 0:
                flash("❌ مبلغ باید بزرگتر از صفر باشد", "error")
                return redirect(url_for('main_bp.cash_initial'))
            
            if not cash:
                cash = CashBalance(amount=0)
                db.session.add(cash)
                db.session.flush()
            
            old_balance = cash.amount
            cash.amount = amount
            cash.updated_by = current_user.id
            cash.last_updated = datetime.now()
            
            # ثبت تراکنش به عنوان موجودی اولیه
            transaction = CashTransaction(
                amount=amount,
                transaction_type='initial',
                description=description or "ثبت موجودی اولیه صندوق",
                balance_before=old_balance,
                balance_after=amount,
                created_by=current_user.id
            )
            db.session.add(transaction)
            db.session.commit()
            
            flash(f"✅ موجودی اولیه صندوق به مبلغ {amount:,.0f} افغانی ثبت شد", "success")
            return redirect(url_for('main_bp.cash_balance'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ خطا: {str(e)}", "error")
            return redirect(url_for('main_bp.cash_initial'))
    
    return render_template("cash_initial.html", now=datetime.now())
@main_bp.route("/cash/debug")
@login_required
def cash_debug():
    """نمایش آخرین تراکنش‌ها برای دیباگ"""
    from app.models import CashTransaction
    
    transactions = CashTransaction.query.order_by(CashTransaction.created_at.desc()).limit(20).all()
    
    result = "<h1>آخرین تراکنش‌ها</h1><table border='1'><tr><th>ID</th><th>نوع</th><th>مبلغ</th><th>توضیحات</th></tr>"
    for t in transactions:
        result += f"<tr><td>{t.id}</td><td>{t.transaction_type}</td><td>{t.amount}</td><td>{t.description}</td></tr>"
    result += "</table>"
    
    return result
@main_bp.route("/export/expenses")
@login_required
def export_expenses():
    """خروجی Excel از مصارف"""
    from datetime import datetime, timedelta
    from app.models import DailyExpense
    from .excel_export import export_to_excel, format_currency
    
    # دریافت پارامترهای فیلتر
    filter_type = request.args.get('filter', 'all')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # ساخت query پایه
    query = DailyExpense.query
    
    # اعمال فیلتر تاریخ
    if filter_type == 'today':
        today = datetime.now().date()
        start_datetime = datetime.combine(today, datetime.min.time())
        end_datetime = datetime.combine(today, datetime.max.time())
        query = query.filter(DailyExpense.date >= start_datetime, 
                            DailyExpense.date <= end_datetime)
    
    elif filter_type == 'week':
        today = datetime.now().date()
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        query = query.filter(DailyExpense.date >= start_datetime,
                            DailyExpense.date <= end_datetime)
    
    elif filter_type == 'month':
        today = datetime.now().date()
        start_date = today.replace(day=1)
        end_date = today
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        query = query.filter(DailyExpense.date >= start_datetime,
                            DailyExpense.date <= end_datetime)
    
    elif start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            start_datetime = datetime.combine(start, datetime.min.time())
            end_datetime = datetime.combine(end, datetime.max.time())
            query = query.filter(DailyExpense.date >= start_datetime,
                                DailyExpense.date <= end_datetime)
        except:
            pass
    
    # دریافت مصارف
    expenses = query.order_by(DailyExpense.date.desc()).all()
    
    # آماده‌سازی داده‌ها
    data = []
    for expense in expenses:
        data.append({
            "شناسه": expense.id,
            "توضیحات": expense.description,
            "مبلغ": format_currency(expense.amount),
            "تاریخ": expense.date.strftime('%Y-%m-%d %H:%M') if expense.date else '---',
            "ثبت‌کننده": expense.user.full_name if expense.user else '---',
            "دسته‌بندی": getattr(expense, 'category', 'سایر')
        })
    
    # محاسبه مجموع
    total_amount = sum(e.amount for e in expenses)
    
    # اضافه کردن ردیف جمع کل
    data.append({
        "شناسه": "",
        "توضیحات": "جمع کل",
        "مبلغ": format_currency(total_amount),
        "تاریخ": "",
        "ثبت‌کننده": "",
        "دسته‌بندی": ""
    })
    
    # ایجاد نام فایل
    if filter_type == 'today':
        filename = f"expenses_{datetime.now().strftime('%Y%m%d')}.xlsx"
    elif filter_type == 'week':
        filename = f"expenses_week_{datetime.now().strftime('%Y%m%d')}.xlsx"
    elif filter_type == 'month':
        filename = f"expenses_month_{datetime.now().strftime('%Y%m')}.xlsx"
    else:
        filename = f"expenses_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return export_to_excel(
        data,
        filename,
        "مصارف",
        f"گزارش مصارف - {filter_type}"
    )
    
@main_bp.route("/notifications")
@login_required
def notifications():
    """صفحه نمایش همه اعلان‌ها"""
    from app.notification_manager import NotificationManager
    
    notifications = NotificationManager.get_all_notifications()
    unread_count = NotificationManager.get_unread_count()
    urgent_count = NotificationManager.get_urgent_count()
    
    return render_template(
        "notifications.html",
        notifications=notifications,
        unread_count=unread_count,
        urgent_count=urgent_count
    )

@main_bp.route("/notifications/mark-read/<int:notification_id>", methods=["POST"])
@login_required
def mark_notification_read(notification_id):
    """علامت‌گذاری اعلان به عنوان خوانده شده"""
    notification = Notification.query.get_or_404(notification_id)
    notification.mark_as_read()
    
    return jsonify({'success': True})

@main_bp.route("/notifications/mark-all-read", methods=["POST"])
@login_required
def mark_all_notifications_read():
    """علامت‌گذاری همه اعلان‌ها به عنوان خوانده شده"""
    Notification.query.filter_by(is_read=False).update({'is_read': True})
    db.session.commit()
    
    flash("✅ همه اعلان‌ها به عنوان خوانده شده علامت‌گذاری شدند", "success")
    return redirect(url_for('main_bp.notifications'))

@main_bp.route("/api/notifications/count")
@login_required
def api_notifications_count():
    """API برای دریافت تعداد اعلان‌ها (برای آیکون منو)"""
    
    
    return jsonify({
        'unread': NotificationManager.get_unread_count(),
        'urgent': NotificationManager.get_urgent_count()
    })
    
    
# در routes.py، این route را اضافه کنید
@main_bp.route("/notifications/refresh")
@login_required
def refresh_notifications():
    """بررسی و به‌روزرسانی اعلان‌ها"""
    from app.notification_manager import NotificationManager
    
    # پاک کردن اعلان‌های قبلی
    Notification.query.delete()
    
    # ایجاد اعلان‌های جدید
    low_count = NotificationManager.check_low_stock()
    expiring_count = NotificationManager.check_expiring_products()
    
    flash(f"✅ {low_count} اعلان موجودی کم و {expiring_count} اعلان نزدیک انقضا ایجاد شد", "success")
    return redirect(url_for('main_bp.notifications'))

@main_bp.route("/creditors/<int:creditor_id>/export")
@login_required
def export_creditor_transactions(creditor_id):
    """خروجی Excel از تراکنش‌های طلبکار"""
    from app.excel_export import export_to_excel, format_currency
    from datetime import datetime
    
    creditor = Creditor.query.get_or_404(creditor_id)
    transactions = DebtTransaction.query.filter_by(creditor_id=creditor.id)\
                    .order_by(DebtTransaction.date_created.desc()).all()
    
    data = []
    for t in transactions:
        data.append({
            "تاریخ": t.date_created.strftime('%Y-%m-%d %H:%M'),
            "نوع تراکنش": "افزایش طلب" if t.transaction_type == 'debt' else "پرداخت",
            "مبلغ (افغانی)": format_currency(t.amount),
            "شماره رسید": t.receipt_number or '---',
            "توضیحات": t.description or '---',
            "ثبت کننده": t.user.full_name if t.user else '---'
        })
    
    # اضافه کردن ردیف جمع کل
    total_debt = sum(t.amount for t in transactions if t.transaction_type == 'debt')
    total_payment = sum(t.amount for t in transactions if t.transaction_type == 'payment')
    
    data.append({
        "تاریخ": "جمع کل",
        "نوع تراکنش": f"مجموع افزایش طلب: {format_currency(total_debt)} - مجموع پرداخت: {format_currency(total_payment)}",
        "مبلغ (افغانی)": "",
        "شماره رسید": "",
        "توضیحات": f"مانده نهایی: {format_currency(creditor.current_debt)} افغانی",
        "ثبت کننده": ""
    })
    
    filename = f"transactions_{creditor.name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return export_to_excel(
        data,
        filename,
        f"تراکنش‌های {creditor.name}",
        f"گزارش تراکنش‌های طلبکار - {creditor.name}"
    )
    # app/routes.py

# ============================================================
# # نمایش تراکنش‌های طلبکار
# # ============================================================
# @main_bp.route('/creditors/<int:creditor_id>/transactions')
# @login_required
# def creditor_transactions(creditor_id):
#     """نمایش لیست تراکنش‌های یک طلبکار"""
#     from app.models import Creditor, DebtTransaction
    
#     creditor = Creditor.query.get_or_404(creditor_id)
    
#     # دریافت تمام تراکنش‌های این طلبکار
#     transactions = DebtTransaction.query.filter_by(creditor_id=creditor_id)\
#                                          .order_by(DebtTransaction.date_created.desc())\
#                                          .all()
    
#     # محاسبه جمع قرض داری و پرداختی‌ها
#     total_debts = sum(t.amount for t in transactions if t.transaction_type == 'debt')
#     total_payments = sum(t.amount for t in transactions if t.transaction_type == 'payment')
#     current_debt = total_debts - total_payments
    
#     return render_template('creditor_transactions.html',
#                          creditor=creditor,
#                          transactions=transactions,
#                          total_debts=total_debts,
#                          total_payments=total_payments,
#                          current_debt=current_debt)
 

# =====================================================
# API برای آپلود عکس داشبورد
# =====================================================
  # این خط را بالای فایل اضافه کنید

# ... بقیه کدها ...
from flask import current_app
@main_bp.route('/api/upload-dashboard-image', methods=['POST'])
@login_required
def upload_dashboard_image():
    try:
        print("🔵 درخواست آپلود دریافت شد")
        
        if 'image' not in request.files:
            return jsonify({'success': False, 'error': 'فایلی ارسال نشده است'}), 400
        
        file = request.files['image']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'فایلی انتخاب نشده است'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'فرمت فایل مجاز نیست'}), 400
        
        # استفاده از مسیر static_folder
        upload_path = os.path.join(current_app.root_path, 'static', 'uploads')
        
        if not os.path.exists(upload_path):
            os.makedirs(upload_path)
        
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"dashboard_image_{uuid.uuid4().hex}.{ext}"
        filepath = os.path.join(upload_path, filename)
        file.save(filepath)
        
        old_image = session.get('dashboard_image')
        if old_image:
            old_path = os.path.join(upload_path, old_image)
            if os.path.exists(old_path):
                os.remove(old_path)
        
        session['dashboard_image'] = filename
        
        return jsonify({'success': True, 'image_url': f"/static/uploads/{filename}"})
        
    except Exception as e:
        print(f"❌ خطا: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    
@main_bp.route('/api/delete-dashboard-image', methods=['DELETE'])
@login_required
def delete_dashboard_image():
    try:
        print("🔵 درخواست حذف عکس دریافت شد")  # برای دیباگ
        
        filename = session.get('dashboard_image')
        print(f"📄 نام فایل در session: {filename}")
        
        if filename:
            upload_path = os.path.join(current_app.root_path, 'static', 'uploads')
            filepath = os.path.join(upload_path, filename)
            
            print(f"📁 مسیر فایل: {filepath}")
            
            if os.path.exists(filepath):
                os.remove(filepath)
                print("✅ فایل حذف شد")
            else:
                print("⚠️ فایل وجود ندارد")
            
            session.pop('dashboard_image', None)
            print("💾 session پاک شد")
            return jsonify({'success': True})
        else:
            print("⚠️ هیچ فایلی در session نیست")
            return jsonify({'success': True, 'message': 'هیچ عکسی برای حذف وجود ندارد'})
        
    except Exception as e:
        print(f"❌ خطا: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
    
@main_bp.route('/bulk_price_update', methods=['GET', 'POST'])
@login_required
def bulk_price_update():

    if request.method == 'POST':

        try:
            operation = request.form.get('operation')
            percent = request.form.get('percent')
            price_type = request.form.get('price_type')

            # بررسی خالی نبودن
            if not percent:
                flash('درصد را وارد کنید', 'danger')
                return redirect(url_for('main_bp.bulk_price_update'))

            percent = float(percent)

            products = Product.query.all()

            for product in products:

                # انتخاب نوع قیمت
                if price_type == 'selling':
                    old_price = product.selling_price or 0
                else:
                    old_price = product.buying_price or 0

                change_amount = old_price * (percent / 100)

                # افزایش یا کاهش
                if operation == 'increase':
                    new_price = old_price + change_amount
                else:
                    new_price = old_price - change_amount

                # جلوگیری از منفی شدن
                if new_price < 0:
                    new_price = 0

                # ذخیره قیمت جدید
                if price_type == 'selling':
                    product.selling_price = round(new_price, 2)
                else:
                    product.buying_price = round(new_price, 2)

            db.session.commit()

            flash('قیمت‌ها با موفقیت تغییر کرد', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'خطا: {str(e)}', 'danger')

        return redirect(url_for('main_bp.bulk_price_update'))

    return render_template('bulk_price_update.html')
@main_bp.route('/product_sales_report', methods=['GET', 'POST'])
@login_required
def product_sales_report():

    products = Product.query.order_by(Product.name).all()

    report_data = []
    selected_product = None

    if request.method == 'POST':

        product_id = request.form.get('product_id')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')

        selected_product = Product.query.get(product_id)

        query = db.session.query(SaleItem).join(Sale)

        query = query.filter(SaleItem.product_id == product_id)

        if start_date:
            query = query.filter(Sale.date >= start_date)

        if end_date:
            query = query.filter(Sale.date <= end_date)

        sale_items = query.all()

        total_qty = 0
        total_sales = 0
        total_profit = 0

        for item in sale_items:

            total_qty += item.quantity
            total_sales += item.final_amount
            total_profit += item.profit or 0

            report_data.append({
                'invoice': item.sale.invoice_number,
                'date': item.sale.date,
                'qty': item.quantity,
                'price': item.selling_price,
                'amount': item.final_amount,
                'profit': item.profit
            })

        summary = {
            'total_qty': total_qty,
            'total_sales': total_sales,
            'total_profit': total_profit
        }

        return render_template(
            'product_sales_report.html',
            products=products,
            report_data=report_data,
            summary=summary,
            selected_product=selected_product
        )

    return render_template(
        'product_sales_report.html',
        products=products,
        report_data=None
    )
    
@main_bp.route("/api/products_list")
def api_products_list():
    """API برای دریافت لیست محصولات (برای جستجوی فروش)"""
    products = Product.query.all()
    return jsonify([{
        'id': p.id,
        'name': p.name,
        'selling_price': p.selling_price,
        'quantity': p.quantity,
        'unit': p.unit,
        'items_per_carton': p.items_per_carton or 1
    } for p in products])
from flask import jsonify, request
from app.models import Product
@main_bp.route('/api/products/search')
@login_required
def search_products():
    query = request.args.get("q", "")
    
    products = Product.query.filter(
        Product.name.ilike(f"%{query}%")
    ).limit(20).all()
    
    return jsonify([
        {
            "id": p.id,
            "name": p.name,
            "batch_no": p.batch_no,  # ✅ اضافه شده
            "selling_price": float(p.selling_price or 0),
            "items_per_carton": p.items_per_carton or 1,
            "total_items": p.total_items,
            "unit": p.unit if p.unit else "عدد",
            "barcode": p.barcode
        }
        for p in products
    ])    
@main_bp.route('/api/customers/add', methods=['POST'])
@login_required
def api_add_customer():
    """API افزودن مشتری جدید (برای فرم فروش)"""
    try:
        # دریافت داده
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'داده ارسال نشده است'}), 400
        
        name = data.get('name', '').strip()
        phone = data.get('phone', '').strip() or None
        address = data.get('address', '').strip() or None
        
        if not name:
            return jsonify({'success': False, 'error': 'نام مشتری الزامی است'}), 400
        
        # ایجاد مشتری جدید
        customer = Customer(
            name=name,
            phone=phone,
            address=address,
            total_debt=0
        )
        db.session.add(customer)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'customer_id': customer.id,
            'customer': {
                'id': customer.id,
                'name': customer.name,
                'phone': customer.phone
            }
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ خطا در افزودن مشتری: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
    
    
@main_bp.route("/debug/check-inventory")
@login_required
def debug_check_inventory():
    """بررسی موجودی واقعی محصولات"""
    products = Product.query.all()
    
    result = []
    for p in products:
        # محاسبه دستی موجودی
        items_per_carton = p.items_per_carton or 1
        manual_total = (p.quantity or 0) * items_per_carton + (p.single_quantity or 0)
        
        result.append({
            'name': p.name,
            'quantity (کارتن)': p.quantity,
            'single_quantity (تک)': p.single_quantity,
            'items_per_carton': items_per_carton,
            'total_items (property)': p.total_items,
            'total_items (محاسبه دستی)': manual_total,
            'selling_price': p.selling_price,
            'buying_price': p.buying_price
        })
    
    # محاسبه مجموع کل
    total_items = sum(r['total_items (محاسبه دستی)'] for r in result)
    total_value = sum(r['total_items (محاسبه دستی)'] * (r['selling_price'] or 0) for r in result)
    
    return {
        'products': result,
        'summary': {
            'total_items': total_items,
            'total_value': total_value,
            'product_count': len(result)
        }
    }
@main_bp.route("/debug/fix-all-inventory")
@login_required
def fix_all_inventory():
    """تصحیح موجودی تمام محصولات - مقادیر Float را گرد می‌کند"""
    products = Product.query.all()
    
    results = []
    for product in products:
        # ذخیره مقادیر قدیم
        old_cartons = product.quantity
        old_singles = product.single_quantity
        
        # اطمینان از نوع داده صحیح
        product.quantity = float(int(product.quantity or 0))  # تبدیل به integer و سپس float
        product.single_quantity = int(product.single_quantity or 0)
        
        # محاسبه موجودی کل
        items_per_carton = float(product.items_per_carton or 1)
        total_items = int(product.quantity * items_per_carton + product.single_quantity)
        
        results.append({
            'id': product.id,
            'name': product.name,
            'old_cartons': old_cartons,
            'new_cartons': product.quantity,
            'old_singles': old_singles,
            'new_singles': product.single_quantity,
            'total_items': total_items,
            'items_per_carton': items_per_carton
        })
    
    db.session.commit()
    
    return {
        'message': 'موجودی‌ها تصحیح شدند',
        'fixed_products': results,
        'total_fixed': len(results)
    }
@main_bp.route("/debug/simple-inventory")
@login_required
def simple_inventory():
    """نمایش ساده موجودی برای تست"""
    products = Product.query.filter(
        (Product.quantity > 0) | (Product.single_quantity > 0)
    ).all()
    
    html = """
    <html dir="rtl">
    <head>
        <style>
            body { font-family: Tahoma; background: #1a1a2e; color: white; padding: 20px; }
            table { width: 100%; border-collapse: collapse; }
            th, td { border: 1px solid #444; padding: 8px; text-align: right; }
            th { background: #0f3460; color: #e94560; }
            tr:nth-child(even) { background: #16213e; }
            .total { margin-top: 20px; padding: 15px; background: #0f3460; border-radius: 10px; }
        </style>
    </head>
    <body>
        <h2>📦 لیست موجودی کالاها</h2>
        <table>
            <tr>
                <th>نام محصول</th>
                <th>کارتن</th>
                <th>تک</th>
                <th>تعداد در کارتن</th>
                <th>کل دانه</th>
                <th>قیمت فروش</th>
                <th>ارزش کل</th>
            </tr>
    """
    
    total_value = 0
    total_items = 0
    
    for p in products:
        items_per_carton = float(p.items_per_carton or 1)
        carton_items = float(p.quantity or 0) * items_per_carton
        single_items = float(p.single_quantity or 0)
        total = int(carton_items + single_items)
        value = total * (p.selling_price or 0)
        
        total_items += total
        total_value += value
        
        html += f"""
            <tr>
                <td>{p.name}</td>
                <td>{p.quantity:.0f}</td>
                <td>{p.single_quantity}</td>
                <td>{p.items_per_carton:.0f}</td>
                <td>{total:,}</td>
                <td>{p.selling_price:,.0f}</td>
                <td>{value:,.0f}</td>
            </tr>
        """
    
    html += f"""
        </table>
        <div class="total">
            <h3>📊 جمع کل:</h3>
            <p>تعداد کل دانه‌ها: {total_items:,}</p>
            <p>ارزش کل موجودی: {total_value:,.0f} افغانی</p>
        </div>
    </body>
    </html>
    """
    return html
@main_bp.route("/cash/export_excel", methods=["GET"])
@login_required
def cash_export_excel():
    """خروجی اکسل از تراکنش‌های صندوق با نمایش ثبت کننده و برداشت کننده"""
    from app.models import CashTransaction, CashWithdrawal
    from datetime import datetime
    import pandas as pd
    from io import BytesIO
    from flask import send_file
    
    try:
        # دریافت تمام تراکنش‌ها
        transactions = CashTransaction.query.order_by(CashTransaction.created_at.desc()).all()
        
        # ایجاد دیکشنری برای نگهداری اطلاعات برداشت‌ها
        withdrawal_info = {}
        withdrawals = CashWithdrawal.query.all()
        for w in withdrawals:
            withdrawal_info[w.id] = {
                'employee_name': w.employee_name if hasattr(w, 'employee_name') else (w.employee.full_name if w.employee else 'نامشخص'),
                'purpose': w.purpose,
                'description': w.description
            }
        
        # آماده سازی داده‌ها برای اکسل
        data = []
        for t in transactions:
            # تعیین نوع تراکنش به فارسی
            if t.transaction_type == 'sale':
                type_fa = 'فروش'
                transaction_detail = ''
                recorder = t.user.full_name if t.user else 'سیستم'
                withdrawer = '-'
                
            elif t.transaction_type == 'expense':
                type_fa = 'مصرف'
                transaction_detail = ''
                recorder = t.user.full_name if t.user else 'سیستم'
                withdrawer = '-'
                
            elif t.transaction_type == 'withdrawal':
                type_fa = 'برداشت'
                # دریافت اطلاعات برداشت از جدول cash_withdrawals
                withdrawal = CashWithdrawal.query.filter_by(id=t.reference_id).first()
                if withdrawal:
                    transaction_detail = f"هدف: {withdrawal.purpose}"
                    # برداشت کننده (کسی که پول را برداشت کرده)
                    if hasattr(withdrawal, 'employee_name') and withdrawal.employee_name:
                        withdrawer = withdrawal.employee_name
                    else:
                        withdrawer = withdrawal.employee.full_name if withdrawal.employee else 'نامشخص'
                    # ثبت کننده (کسی که در سیستم ثبت کرده - معمولاً کاربر فعلی)
                    recorder = t.user.full_name if t.user else 'سیستم'
                else:
                    transaction_detail = ''
                    withdrawer = 'نامشخص'
                    recorder = t.user.full_name if t.user else 'سیستم'
                    
            elif t.transaction_type == 'initial':
                type_fa = 'موجودی اولیه'
                transaction_detail = ''
                recorder = t.user.full_name if t.user else 'سیستم'
                withdrawer = '-'
                
            elif t.transaction_type == 'deposit':
                type_fa = 'واریز'
                transaction_detail = ''
                recorder = t.user.full_name if t.user else 'سیستم'
                withdrawer = '-'
                
            elif t.transaction_type == 'sale_refund':
                type_fa = 'برگشت از فروش'
                transaction_detail = ''
                recorder = t.user.full_name if t.user else 'سیستم'
                withdrawer = '-'
                
            else:
                type_fa = t.transaction_type
                transaction_detail = ''
                recorder = t.user.full_name if t.user else 'سیستم'
                withdrawer = '-'
            
            data.append({
                'ردیف': t.id,
                'تاریخ و ساعت': t.created_at.strftime('%Y/%m/%d %H:%M:%S') if t.created_at else '',
                'نوع تراکنش': type_fa,
                'توضیحات': t.description or transaction_detail,
                'مبلغ (افغانی)': t.amount,
                'موجودی قبل': t.balance_before,
                'موجودی بعد': t.balance_after,
                'ثبت کننده (کاربر سیستم)': recorder,
                'برداشت کننده (گیرنده وجه)': withdrawer
            })
        
        # ایجاد دیتافریم pandas
        df = pd.DataFrame(data)
        
        # محاسبه جمع‌ها
        total_income = df[df['مبلغ (افغانی)'] > 0]['مبلغ (افغانی)'].sum()
        total_expense = df[df['مبلغ (افغانی)'] < 0]['مبلغ (افغانی)'].sum()
        
        # اضافه کردن ردیف جمع‌بندی
        summary = pd.DataFrame([{
            'ردیف': '',
            'تاریخ و ساعت': 'جمع کل',
            'نوع تراکنش': '',
            'توضیحات': f'جمع دریافتی‌ها: {total_income:,.0f} | جمع پرداختی‌ها: {abs(total_expense):,.0f} | مانده: {total_income - abs(total_expense):,.0f}',
            'مبلغ (افغانی)': '',
            'موجودی قبل': '',
            'موجودی بعد': '',
            'ثبت کننده (کاربر سیستم)': '',
            'برداشت کننده (گیرنده وجه)': ''
        }])
        
        df = pd.concat([df, summary], ignore_index=True)
        
        # ایجاد فایل اکسل
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='تراکنش‌های صندوق', index=False)
            
            # تنظیم عرض ستون‌ها
            worksheet = writer.sheets['تراکنش‌های صندوق']
            
            # تنظیم رنگ و استایل برای هدر
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # تنظیم عرض ستون‌ها
            column_widths = {
                'A': 10,  # ردیف
                'B': 20,  # تاریخ و ساعت
                'C': 15,  # نوع تراکنش
                'D': 40,  # توضیحات
                'E': 15,  # مبلغ
                'F': 15,  # موجودی قبل
                'G': 15,  # موجودی بعد
                'H': 20,  # ثبت کننده
                'I': 20   # برداشت کننده
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # رنگ‌آمیزی ردیف‌ها بر اساس نوع تراکنش
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=3)  # ستون نوع تراکنش
                if cell.value == 'برداشت':
                    for col in range(1, 10):
                        cell_color = worksheet.cell(row=row, column=col)
                        cell_color.fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                elif cell.value == 'فروش':
                    for col in range(1, 10):
                        cell_color = worksheet.cell(row=row, column=col)
                        cell_color.fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
        
        output.seek(0)
        
        # نام فایل با تاریخ
        filename = f"cash_transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except ImportError as e:
        flash(f"❌ کتابخانه مورد نیاز نصب نیست. لطفاً نصب کنید: pip install pandas openpyxl", "danger")
        print(f"Import Error: {e}")
        return redirect(url_for("main_bp.cash_balance"))
    except Exception as e:
        flash(f"❌ خطا در ایجاد فایل اکسل: {str(e)}", "danger")
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return redirect(url_for("main_bp.cash_balance"))
    
@main_bp.route("/cash/withdrawals_export_excel", methods=["GET"])
@login_required
def cash_withdrawals_export_excel():
    """خروجی اکسل از برداشت‌های صندوق"""
    from app.models import CashWithdrawal
    from datetime import datetime
    import pandas as pd
    from io import BytesIO
    from flask import send_file
    
    try:
        withdrawals = CashWithdrawal.query.order_by(CashWithdrawal.withdrawal_date.desc()).all()
        
        data = []
        for w in withdrawals:
            data.append({
                'ردیف': w.id,
                'تاریخ برداشت': w.withdrawal_date.strftime('%Y/%m/%d %H:%M:%S') if w.withdrawal_date else '',
                'برداشت کننده': w.employee_name if hasattr(w, 'employee_name') else (w.employee.full_name if w.employee else 'نامشخص'),
                'هدف برداشت': w.purpose,
                'مبلغ (افغانی)': w.amount,
                'توضیحات': w.description or '',
                'تاریخ ثبت': w.created_at.strftime('%Y/%m/%d %H:%M:%S') if w.created_at else ''
            })
        
        df = pd.DataFrame(data)
        
        # محاسبه جمع کل
        total_withdrawals = df['مبلغ (افغانی)'].sum()
        
        # اضافه کردن ردیف جمع‌بندی
        summary = pd.DataFrame([{
            'ردیف': '',
            'تاریخ برداشت': 'جمع کل',
            'برداشت کننده': '',
            'هدف برداشت': '',
            'مبلغ (افغانی)': total_withdrawals,
            'توضیحات': '',
            'تاریخ ثبت': ''
        }])
        
        df = pd.concat([df, summary], ignore_index=True)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='برداشت‌های صندوق', index=False)
            
            worksheet = writer.sheets['برداشت‌های صندوق']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        filename = f"cash_withdrawals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f"❌ خطا در ایجاد فایل اکسل: {str(e)}", "danger")
        return redirect(url_for("main_bp.cash_balance"))