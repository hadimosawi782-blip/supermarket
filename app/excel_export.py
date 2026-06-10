# app/excel_export.py
import pandas as pd
from io import BytesIO
from flask import send_file
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel(data, filename, sheet_name="Sheet1", title=None):
    """تابع اصلی برای خروجی Excel"""
    try:
        # ایجاد DataFrame
        df = pd.DataFrame(data)
        
        # اگر داده خالی است
        if df.empty:
            df = pd.DataFrame([{"پیام": "داده‌ای برای نمایش وجود ندارد"}])
        
        # ایجاد فایل Excel در حافظه
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2 if title else 0)
            
            # دریافت workbook و worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # اعمال استایل به هدر
            if title:
                try:
                    last_col_letter = get_column_letter(len(df.columns))
                    merge_range = f'A1:{last_col_letter}1'
                    
                    worksheet.merge_cells(merge_range)
                    title_cell = worksheet['A1']
                    title_cell.value = title
                    title_cell.font = Font(size=14, bold=True)
                    title_cell.alignment = Alignment(horizontal='center')
                    
                    worksheet['A2'] = f"تاریخ ایجاد: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    worksheet['A2'].font = Font(size=10, italic=True)
                except Exception as e:
                    print(f"خطا در ایجاد عنوان: {e}")
            
            # استایل هدر ستون‌ها
            header_row = 3 if title else 1
            for col_idx, col_name in enumerate(df.columns, 1):
                try:
                    cell = worksheet.cell(row=header_row, column=col_idx)
                    cell.value = col_name
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                except Exception as e:
                    print(f"خطا در استایل ستون {col_idx}: {e}")
            
            # تنظیم عرض ستون‌ها
            for col_idx, col in enumerate(df.columns, 1):
                try:
                    max_length = max(
                        df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                        len(str(col))
                    ) + 2
                    column_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[column_letter].width = min(max_length, 50)
                except Exception as e:
                    print(f"خطا در تنظیم عرض ستون {col_idx}: {e}")
            
            # اگر ردیف آخر جمع کل است، آن را پررنگ کن
            if len(df) > 0 and 'توضیحات' in df.columns and df.iloc[-1]['توضیحات'] == 'جمع کل':
                last_row = len(df) + header_row
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=last_row, column=col_idx)
                    cell.font = Font(bold=True)
        
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"خطای کلی در export_to_excel: {e}")
        # برگرداندن یک فایل ساده در صورت خطا
        import io
        simple_output = io.BytesIO()
        simple_df = pd.DataFrame([{"خطا": str(e)}])
        simple_df.to_excel(simple_output, index=False)
        simple_output.seek(0)
        return send_file(
            simple_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"error_{filename}"
        )

def format_currency(value):
    """تبدیل عدد به فرمت پول با جداکننده هزارگان"""
    try:
        if value is None:
            return "0 افغانی"
        return f"{float(value):,.0f} افغانی"
    except (ValueError, TypeError):
        return str(value) if value else "0 افغانی"

def format_number(value):
    """تبدیل عدد به فرمت با جداکننده هزارگان (بدون واحد)"""
    try:
        if value is None:
            return "0"
        return f"{float(value):,.0f}"
    except (ValueError, TypeError):
        return str(value) if value else "0"

def format_percentage(value):
    """تبدیل به درصد"""
    try:
        if value is None:
            return "0%"
        return f"{float(value):.1f}%"
    except (ValueError, TypeError):
        return "0%"